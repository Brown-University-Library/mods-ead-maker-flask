import os, csv, xlsxwriter
from lxml import etree
from openpyxl import load_workbook
import openpyxl
from lxml.builder import ElementMaker
import string, codecs
import chardet
import datetime
import re
import xlrd
from copy import copy
import sys
import traceback


requiredcolumns = ["barcode","subgroupTitle", "subgroupID", "recordgroupTitle", "recordgroupID", "locationCopies", "subjectTopicsFAST","Ignore", "seriesTitle", "seriesID", "subSeriesTitle", "subSeriesID", "fileTitle", "itemTitle", "subTitle", "place","dateText","dateStart","dateEnd","dateBulkStart","dateBulkEnd","dateQualifier", "shelfLocator1", "shelfLocator1ID", "shelfLocator2", "shelfLocator2ID", "shelfLocator3","shelfLocator3ID","typeOfResource","genreAAT","genreLCSH","genreLocal","genreRBGENR","extentQuantity","extentSize","extentSpeed","form","noteScope","noteHistorical","noteHistoricalClassYear","noteGeneral","language","noteAccession","identifierBDR","publisher","namePersonCreatorLC","namePersonCreatorLocal","nameCorpCreatorLC","nameCorpCreatorLocal","namePersonOtherLC","namePersonOtherLocal","subjectNamesLC","subjectNamesLocal","subjectCorpLC","subjectCorpLocal","subjectTopicsLC","subjectTopicsLocal","subjectGeoLC","subjectTemporalLC","subjectTitleLC","collection","dateTextParent","callNumber","repository","findingAid","digitalOrigin","rightsStatementText","rightsStatementURI", "useAndReproduction", "coordinates", "scale", "projection", "containerSummary"]
langcode = {}
langcodeopp = {}
scriptcode = {}
langissue = False

def getSplitCharacter(string):
    if ";" in string:
        return(";")
    else:
        return("|")

def messageToUser(messagetitle, message):
    print("", file=sys.stderr)
    print(messagetitle, file=sys.stderr)
    print(message, file=sys.stderr)
    try:
        raw_input("Press Enter to continue . . .")
    except SyntaxError:
        print("Syntax Error", file=sys.stderr)
    except TypeError:
        print("Type Error", file=sys.stderr)

def multilinefield(refdict, parentelement, originalfieldname, eadfieldname):
    newelement = etree.SubElement(parentelement, eadfieldname)
    lines = refdict.get(originalfieldname, '').splitlines()
    for line in lines:
        pelement = etree.SubElement(newelement, "p")
        pelement.text = ' '.join(line.split())

def repeatingsubjectfield(parentelement, refdict, originalfieldname, eadfieldname, eadattributes):
    splitcharacter = ";"

    for namesindex, addedentry in enumerate(refdict.get(originalfieldname, '').split(splitcharacter)):

        customAttributes = eadattributes.copy()

        #Extract URI

        uri = re.findall("(?P<url>https?://[^\s]+)", addedentry)

        #If there's a URI
        if len(uri) > 0:
            #Remove it from the addedentry
            addedentry = addedentry.replace(uri[0],"")
            #Add it as a valueURI attribute
            customAttributes["authfilenumber"] = xmltext(uri[0])

        namecontrolaccesselement = etree.SubElement(parentelement, eadfieldname, customAttributes)
        namecontrolaccesselement.text = ' '.join(addedentry.replace("|d", "").replace("|e", "").split())

def repeatingNameField(parentElement, elementName, rowString, assignedRole, source):
    for name in rowString.split(';'):
        currentname = ""
        currentrole = ""
        attributes = {}

        #Extract URI

        uri = re.findall("(?P<url>https?://[^\s]+)", name)

        #If there's a URI
        if len(uri) > 0:
            #Remove it from the addedentry
            name = name.replace(uri[0],"")
            #Add it as a valueURI attribute
            attributes["authfilenumber"] = xmltext(uri[0])

        for index, namefield in enumerate(name.split(',')):
            namefieldrevised = xmltext(namefield)

            if index == 0:
                currentname = currentname + namefieldrevised + ", "
            elif hasYear(namefieldrevised) == True:
                currentname = currentname + namefieldrevised
            elif isAllLower(namefieldrevised) == True:
                currentrole = namefieldrevised
            elif hasLetters(namefieldrevised) != None:
                currentname = currentname + namefieldrevised + ", "

        if currentrole:
            attributes['role'] = currentrole
        elif assignedRole:
            attributes['role'] = assignedRole

        attributes['source'] = source

        nameelement = etree.SubElement(parentElement, elementName, attributes)
        nameelement.text = xmltext(currentname).rstrip(',').lstrip(',')

def xmltext(text):
    return(' '.join(str(text).split()))

def copyworkbook(path1, path2):

    wb1 = load_workbook(filename=path1)
    ws1 = wb1.worksheets[0]

    wb2 = load_workbook(filename=path2)
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value
            if cell.has_style:
                ws2[cell.coordinate].font = copy(cell.font)
                ws2[cell.coordinate].border = copy(cell.border)
                ws2[cell.coordinate].fill = copy(cell.fill)
                ws2[cell.coordinate].number_format = copy(cell.number_format)
                ws2[cell.coordinate].protection = copy(cell.protection)
                ws2[cell.coordinate].alignment = copy(cell.alignment)

    wb2.save(path2)

def convertEncoding(from_encode,to_encode,old_filepath,target_file):
    f1=open(old_filepath)
    content2=[]
    while True:
        line=f1.readline()
        content2.append(line.decode(from_encode).encode(to_encode))
        if len(line) ==0:
            break

    f1.close()
    f2=open(target_file,'w')
    f2.writelines(content2)
    f2.close()

def hasNumbers(s):
    return any(i.isdigit() for i in s)

def hasLetters(s):
    return re.search('[a-zA-Z]', s)

def hasYear(s):
    numbercount = 0
    for i in s:
        if i.isdigit() == True:
            numbercount = numbercount + 1
    if numbercount > 3:
        return True
    else:
        return False

def isAllLower(s):
    nonlowercase = 0
    for i in s.replace(' ', ''):
        if i.islower() == False:
            nonlowercase = nonlowercase + 1
            break
    if nonlowercase > 0:
        return False
    else:
        return True



def let_user_pick(message, options):
    print("", file=sys.stderr)
    print(message, file=sys.stderr)
    for idx, element in enumerate(options):
        print("{}) {}".format(idx+1,element), file=sys.stderr)
    i = input("Enter number: ")
    try:
        if 0 < int(i) <= len(options):
            return options[i-1]
    except:
        pass
    return None

def XLSDictReader(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        rowarray = []

        for row in range(1, sheet.nrows):
            rowdictionary = {}
            for column in range(sheet.ncols):
                #If the value is a number, turn it into a string.
                newvalue = ''
                if sheet.cell(row,column).ctype > 1:
                    newvalue = str(sheet.cell_value(row,column)).replace('|d', '').replace('|e', '').replace('|',';')
                    #print(sheet.cell_value(row,column))
                else:
                    newvalue = sheet.cell_value(row,column).replace('|d', '').replace('|e', '').replace('|',';')

                #If the column is repeating, serialize the row values.
                if rowdictionary.get(sheet.cell_value(0,column), '') != '':
                    rowdictionary[sheet.cell_value(0,column)] = rowdictionary[sheet.cell_value(0,column)] + ';' + newvalue
                else:
                    rowdictionary[sheet.cell_value(0,column)] = newvalue
            rowarray.append(rowdictionary)
        return(rowarray)

def XLSDictReaderVertical(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        verticaldictionary = {}
        datacolumn = 3

        for column in range(0,sheet.ncols):
            #print(sheet.cell_value(0, column))
            if "Data Entry" in sheet.cell_value(0, column):
                datacolumn = column

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 0)
            value = ""
            if sheet.cell(row,datacolumn).ctype > 1:
                    value = str(sheet.cell_value(row,datacolumn)).replace('|d', '').replace('|e', '').replace('|',';')
            else:
                    value = sheet.cell_value(row,datacolumn).replace('|d', '').replace('|e', '').replace('|',';')
            if verticaldictionary.get(key, '') == '':
                verticaldictionary[key] = value
            else:
                verticaldictionary[key] = verticaldictionary[key] + ';' + value
        return(verticaldictionary)

def XLSDictReaderLanguageCode(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        langcode = {}

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 0)
            value = sheet.cell_value(row, 1)
            langcode[key] = value
        return(langcode)

def XLSDictReaderLanguageCodeOpp(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        langcode = {}

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 1)
            value = sheet.cell_value(row, 0)
            langcode[key] = value
        return(langcode)

def XLSDictReaderScriptCode(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        scriptcode = {}

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 0)
            value = sheet.cell_value(row, 2)
            scriptcode[key] = value
        return(scriptcode)

def getSheetNames(chosenfile):
    excel = xlrd.open_workbook(chosenfile)
    sheetnames = excel.sheet_names()
    return(sheetnames)

CACHEDIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache") + "/"
#CACHEDIR = os.path.dirname(os.path.abspath(__file__)) + "/"
HOMEDIR = os.path.dirname(os.path.abspath(__file__)) + "/"
#HOMEDIR = os.path.dirname(os.path.abspath(__file__)) + "/"

print("._. EAD Maker ._.", file=sys.stderr)

def processExceltoEAD(chosenfile, chosensheet, id):

    if not os.path.exists(CACHEDIR + id):
            os.mkdir(CACHEDIR + id)

    #Get all languages codes and script codes.
    langcode = XLSDictReaderLanguageCode(HOMEDIR + "SupportedLanguages.xlsx","languages xlsx")
    langcodeopp = XLSDictReaderLanguageCodeOpp(HOMEDIR + "SupportedLanguages.xlsx","languages xlsx")
    scriptcode = XLSDictReaderScriptCode(HOMEDIR + "SupportedLanguages.xlsx","languages xlsx")

    csvdata = {}
    cldata = {}
    langissue = False

    excel = xlrd.open_workbook(chosenfile)
    sheetnames = excel.sheet_names()
    selectedsheet = excel.sheet_by_name(chosensheet)
    columnsinsheet = [str(cell.value) for cell in selectedsheet.row(0)]

    missingcolumns = []
    for column in requiredcolumns:
        if (column in columnsinsheet) == False:
            #print("Missing spreadsheet column: " + column + '\n')
            missingcolumns.append(column)

    if len(missingcolumns) != 0:
        print("*Missing Columns Detected*" + '\n', file=sys.stderr)
        print("The columns below are missing from your spreadsheet. The script will continue without them." + '\n\n', file=sys.stderr)

        for column in missingcolumns:
            print("   " + column + '\n', file=sys.stderr)


    print('\n\n', file=sys.stderr)


    csvdata = XLSDictReader(chosenfile, chosensheet)

    if "Collection-Level Data" not in sheetnames:
        copyworkbook(HOMEDIR + "Collection-Level Data.xlsx", chosenfile)
        excel = xlrd.open_workbook(chosenfile)

        #if originalfile != '':
        #    copyworkbook(os.path.dirname(os.path.abspath(__file__)) + "/data/Collection-Level Data.xlsx", originalfile)

        #print("")
        print("*Collection-Level Data Missing*\n", file=sys.stderr)
        print("Collection-level data is missing from your spreadsheet. A sheet titled Collection-Level Data has been automatically added. Enter data in this sheet to add collection-level data to your EAD file.\n", file=sys.stderr)


        print('\n\n', file=sys.stderr)


        #try:
        #    raw_input("Press Enter to continue . . . ")
        #except SyntaxError:
        #    raw_input = 0
        #except TypeError:
        #    raw_input = 0

    cldata = XLSDictReaderVertical(chosenfile, "Collection-Level Data")
    chosenfile = chosensheet

    #Create the output directory and save the path to the output_path variable.
    now = datetime.datetime.now()

    #try:
   #      os.mkdir(output_path + '/'+ chosenfile + " " + now.strftime("%m-%d-%Y %H %M " + str(now.second)))
   # except OSError:
   #      print ("")
   # else:
#    print ("")
   #    output_path = output_path + '/'+ chosenfile + " " + now.strftime("%m-%d-%Y %H %M " + str(now.second))

    #Create the error CSV.
    #errorfile = open(output_path + '/Error Report ' + now.strftime("%m-%d-%Y %H %M " + str(now.second)) + '.csv', mode='wb')
    #errorcsvwriter = csv.writer(errorfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    #errorcsvwriter.writerow(['Spreadsheet Row', 'BDR Number', 'Column Name', 'Column Contents', 'Potential Issue'])

    #Set up namespaces and attributes for XML.
    attr_qname = etree.QName("http://www.w3.org/2001/XMLSchema-instance", "schemaLocation")
    ns_map = {"ead":"urn:isbn:1-931666-22-9", "ns2" : "http://www.w3.org/1999/xlink", "xsi" : "http://www.w3.org/2001/XMLSchema-instance"}
    ns_map2 = {"ns2":"http://www.w3.org/1999/xlink"}
    #Create top elements for EAD.
    eadtop = etree.Element("ead", {attr_qname: "urn:isbn:1-931666-22-9 http://www.loc.gov/ead/ead.xsd", "audience":"external","relatedencoding":"MARC21", "xmlns":"urn:isbn:1-931666-22-9"}, nsmap=ns_map)
    eadheaderelement = etree.SubElement(eadtop, "eadheader", {"audience":"external","countryencoding":"iso3166-1","dateencoding":"iso8601","scriptencoding":"iso15924", "relatedencoding":"MARC21", "repositoryencoding":"iso15511","langencoding":"iso639-2b"})

    #Create archival description elements.
    archdescelement = etree.SubElement(eadtop, "archdesc", {"level":"collection", "type":"inventory"})
    coldidelement = etree.SubElement(archdescelement, "did")

    #Create dsc element and the ctelement variable, which will hold the current series or subseries the script is adding files/items to.
    cseriesID = ""
    ccontrolaccess = etree.Element("ignore")
    cdid = etree.Element("ignoreagain")
    cunittitle = etree.Element("ignoreagainagain")
    cserieslist = etree.Element("continuetoignore")
    csubserieslist = etree.Element("ignoreagainagain")

    #Collection-level description in the archdesc element.
    primaryunittitle = etree.SubElement(coldidelement, "unittitle", {"type":"primary"})
    primaryunittitle.text = xmltext(cldata.get("title", ''))

    filingunittitle = etree.SubElement(coldidelement, "unittitle", {"type":"filing"})
    filingunittitle.text = xmltext(cldata.get("filingTitle", ''))

    colunitidelement = etree.SubElement(coldidelement, "unitid", {"countrycode":"US","repositorycode":"US-"+xmltext(cldata.get("MARCRepositoryCode", '')),"type":"collection"})
    colunitidelement.text = xmltext(cldata.get("callNumber", ''))

    colrepositoryelement = etree.SubElement(coldidelement, "repository")
    repositorycorpelement = etree.SubElement(colrepositoryelement, "corpname")
    repositorycorpelement.text = xmltext(cldata.get("repositoryCorporateName", ''))
    repositorysubarea = etree.SubElement(repositorycorpelement, "subarea")
    repositorysubarea.text = xmltext(cldata.get("repositoryCorporateSubarea", ''))

    coladdresselement = etree.SubElement(colrepositoryelement, "address")
    coladdresslines = cldata.get("repositoryAddress", '').splitlines()
    for line in coladdresslines:
        addresselement = etree.SubElement(coladdresselement, "addressline")
        addresselement.text = xmltext(line)

    collangmaterial = etree.SubElement(coldidelement, "langmaterial")
    collangmaterialsplitchar = getSplitCharacter(cldata.get("materialLanguage", ''))

    for language in cldata.get("materialLanguage", '').split(collangmaterialsplitchar):
        if language == "":
             continue
        if xmltext(language) in langcode:
             langusagelangelement = etree.SubElement(collangmaterial, "language", {"langcode":langcode[xmltext(language)], "scriptcode":scriptcode[xmltext(language)]})
             langusagelangelement.text =  xmltext(language)

             if scriptcode == "N/A":
                 langissue = True
        else:
             langusagelangelement = etree.SubElement(collangmaterial, "language", {"langcode":"***", "scriptcode":"***"})
             langusagelangelement.text =  xmltext(language)

             langissue = True

    colphysdescelement = etree.SubElement(coldidelement, "physdesc")
    colextentelement = etree.SubElement(colphysdescelement, "extent").text = xmltext(cldata.get("sizeExtent", ''))

    inclusivedateattributes = {"type":"inclusive", "era":"ce","calendar":"gregorian","normal":xmltext(cldata.get("inclusiveDates", '')).replace('-','/').replace(' ','')}
    inclusivedateelement = etree.SubElement(coldidelement, "unitdate", inclusivedateattributes)
    inclusivedateelement.text = xmltext(cldata.get("inclusiveDates", '').replace(' ',''))

    bulkdateattributes = {"type":"bulk", "era":"ce","calendar":"gregorian","normal":xmltext(cldata.get("bulkDates", '')).replace('-','/').replace(' ','')}
    bulkdateelement = etree.SubElement(coldidelement, "unitdate", bulkdateattributes)
    bulkdateelement.text = "(bulk "+ xmltext(cldata.get("bulkDates", '').replace(' ','')) + ")"

    coloriginationelement = etree.SubElement(coldidelement, "origination", {"label":"creator"})

    colpersoncreatorelement = etree.SubElement(coloriginationelement, "persname", {"role":"creator"})
    colpersoncreatorelement.text = xmltext(cldata.get("creatorPerson", '')).replace('|d','')

    colcorporatecreatorelement = etree.SubElement(coloriginationelement, "corpname", {"role":"creator"})
    colcorporatecreatorelement.text = xmltext(cldata.get("creatorCorporate", ''))

    colabstractelement = etree.SubElement(coldidelement, "abstract")
    colabstractelement.text = xmltext(cldata.get("abstract", ''))

    #After the collection's did element.
    colbioghistelement = etree.SubElement(archdescelement, "bioghist")
    colbioghistheadelement = etree.SubElement(colbioghistelement, "head")
    colbioghistheadelement.text = "Biographical/Historical Note"

    bioghistlines = cldata.get("bioHistNote", '').splitlines()
    for line in bioghistlines:
        pelement = etree.SubElement(colbioghistelement, "p")
        pelement.text = ' '.join(line.split())

    #descriptive descgrp
    coldescgrpdescriptiveelement = etree.SubElement(archdescelement, "descgrp", {"type":"descriptive"})

    colcollectioninformationelement = etree.SubElement(coldescgrpdescriptiveelement, "head")
    colcollectioninformationelement.text = "Collection information"

    colscopenoteelement = etree.SubElement(coldescgrpdescriptiveelement, "scopecontent")

    colscopelines = cldata.get("scopeNote", '').splitlines()
    for line in colscopelines:
        pelement = etree.SubElement(colscopenoteelement, "p")
        pelement.text = ' '.join(line.split())

    coluserestrictelement = etree.SubElement(coldescgrpdescriptiveelement, "userestrict")
    coluserestrictlines = cldata.get("conditionsUse", '').splitlines()
    for line in coluserestrictlines:
        pelement = etree.SubElement(coluserestrictelement, "p")
        pelement.text = ' '.join(line.split())

    colaccessrestrictelement = etree.SubElement(coldescgrpdescriptiveelement, "accessrestrict")
    colaccessrestrictlines = cldata.get("conditionsAccess", '').splitlines()
    for line in colaccessrestrictlines:
        pelement = etree.SubElement(colaccessrestrictelement, "p")
        pelement.text = ' '.join(line.split())

    colpreferciteelement = etree.SubElement(coldescgrpdescriptiveelement, "prefercite")
    colprefercitelines = cldata.get("preferredCitation", '').splitlines()
    for line in colprefercitelines:
        pelement = etree.SubElement(colpreferciteelement, "p")
        pelement.text = ' '.join(line.split())

    colarrangementelement = etree.SubElement(coldescgrpdescriptiveelement, "arrangement")
    arrangementnotepelement = etree.SubElement(colarrangementelement, "p")
    arrangementnotepelement.text = xmltext(cldata.get("arrangementNote", ''))
    conlyserieslist = etree.SubElement(colarrangementelement, "list")

    #administrative descgrp
    coldescgrpadministrativeelement = etree.SubElement(archdescelement, "descgrp", {"type":"administrative"})

    administrativeinformationcodes = ['acquisitionInformation',"processingInformation","custodialHistory",'accruals','appraisal']

    for code in administrativeinformationcodes:
        if cldata.get(code,'') != '':
            coladministrativeheadelement = etree.SubElement(coldescgrpadministrativeelement, "head")
            coladministrativeheadelement.text = "Administrative information"
            break

    colacqinfoadminelement = etree.SubElement(coldescgrpadministrativeelement, "acqinfo")
    colacqinfoadminlines = cldata.get("acquisitionInformation", '').splitlines()
    for line in colacqinfoadminlines:
        pelement = etree.SubElement(colacqinfoadminelement, "p")
        pelement.text = ' '.join(line.split())

    colprocessinfoadminelement = etree.SubElement(coldescgrpadministrativeelement, "processinfo")
    colprocessinfoadminlines = cldata.get("processingInformation", '').splitlines()
    for line in colprocessinfoadminlines:
        pelement = etree.SubElement(colprocessinfoadminelement, "p")
        pelement.text = ' '.join(line.split())

    colcustodhistadminelement = etree.SubElement(coldescgrpadministrativeelement, "custodhist")
    colcustodhistadminlines = cldata.get("custodialHistory", '').splitlines()
    for line in colcustodhistadminlines:
        pelement = etree.SubElement(colcustodhistadminelement, "p")
        pelement.text = ' '.join(line.split())

    multilinefield(cldata, coldescgrpadministrativeelement, 'accruals', 'accruals')
    multilinefield(cldata, coldescgrpadministrativeelement, 'appraisal', 'appraisal')

    #descgrp additional
    coldescgrpadditionalelement = etree.SubElement(archdescelement, "descgrp", {"type":"additional"})

    if cldata.get('generalNote','') != '' or cldata.get('relatedMaterials','') != '' or cldata.get('separatedMaterials','') != '' or cldata.get('locationOriginals','') != '' or cldata.get('otherFindingAids','') != '' or cldata.get('otherFormats','') != '':
        coldescgrpadditionalheaderelement = etree.SubElement(coldescgrpadditionalelement, "head")
        coldescgrpadditionalheaderelement.text = "Additional information"

    multilinefield(cldata, coldescgrpadditionalelement, 'generalNote', 'odd')
    multilinefield(cldata, coldescgrpadditionalelement, 'relatedMaterials', 'relatedmaterial')
    multilinefield(cldata, coldescgrpadditionalelement, 'separatedMaterials', 'separatedmaterial')
    multilinefield(cldata, coldescgrpadditionalelement, 'locationOriginals', 'originalsloc')
    multilinefield(cldata, coldescgrpadditionalelement, 'otherFindingAids', 'otherfindaid')
    multilinefield(cldata, coldescgrpadditionalelement, 'otherFormats', 'altformavail')

    #descgrp cataloging
    coldescgrpcatalogingelement = etree.SubElement(archdescelement, "descgrp", {"type":"cataloging"})
    coldescgrpnamescontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")

    coldescgrpcontrolaccessnamesheadelement = etree.SubElement(coldescgrpnamescontrolaccesselement, "head")
    if cldata.get("addedEntryPersonLC", '') != '' or cldata.get("addedEntryPersonLocal", '') != '' or cldata.get("addedEntryCorporateLC", '') != '' or cldata.get("addedEntryCorporateLocal", '') != '':
    	coldescgrpcontrolaccessnamesheadelement.text = "Names"

    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryPersonLC', 'persname', {'source':'lcnaf'})
    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryPersonLocal', 'persname', {'source':'local'})

    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryCorporateLC', 'corpname', {'source':'lcnaf'})
    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryCorporateLocal', 'corpname', {'source':'local'})

    coldescgrpsubjectscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccesssubjectheadelement = etree.SubElement(coldescgrpsubjectscontrolaccesselement, "head")
    if cldata.get("addedEntrySubjectLC", '') != '' or cldata.get("addedEntrySubjectLocal", '') != '' or cldata.get("addedEntrySubjectFAST", '') != '' or cldata.get("addedEntryGeographicLC", '') != '' or cldata.get("addedEntryGeographicLocal", '') != '':
    	coldescgrpcontrolaccesssubjectheadelement.text = "Subjects"

    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntrySubjectLC', 'subject', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntrySubjectLocal', 'subject', {'source':'local'})
    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntrySubjectFAST', 'subject', {'source':'fast'})

    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntryGeographicLC', 'geogname', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntryGeographicLocal', 'geogname', {'source':'local'})

    coldescgrpoccupationscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccessoccupationheadelement = etree.SubElement(coldescgrpoccupationscontrolaccesselement, "head")
    if cldata.get("addedEntryOccupationLC", '') != '' or cldata.get("addedEntryOccupationLocal", '') != '':
    	coldescgrpcontrolaccessoccupationheadelement.text = "Occupations"

    repeatingsubjectfield(coldescgrpoccupationscontrolaccesselement, cldata, 'addedEntryOccupationLC', 'occupation', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpoccupationscontrolaccesselement, cldata, 'addedEntryOccupationLocal', 'occupation', {'source':'local'})

    coldescgrpmaterialscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccessgenreheadelement = etree.SubElement(coldescgrpmaterialscontrolaccesselement, "head")
    if cldata.get("addedEntryGenreAAT", '') != '' or cldata.get("addedEntryGenreLCSH", '') != '' or cldata.get("addedEntryGenreTGM", '') != '' or cldata.get("addedEntryGenreRBGENR", '') != '' or cldata.get("addedEntryGenreLocal", '') != '':
    	coldescgrpcontrolaccessgenreheadelement.text = "Types of Materials"

    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreAAT', 'genreform', {'source':'aat'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreLCSH', 'genreform', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreTGM', 'genreform', {'source':'tgm'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreRBGENR', 'genreform', {'source':'rbgenr'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreLocal', 'genreform', {'source':'local'})

    coldescgrptitlescontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccesstitleheadelement = etree.SubElement(coldescgrptitlescontrolaccesselement, "head")
    if cldata.get("addedEntryTitle", '') != '':
    	coldescgrpcontrolaccesstitleheadelement.text = "Titles"

    repeatingsubjectfield(coldescgrptitlescontrolaccesselement, cldata, 'addedEntryTitle', 'title', {})

    coldescgrpriamcoscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccessriamcoheadelement = etree.SubElement(coldescgrpriamcoscontrolaccesselement, "head")
    if cldata.get("RIAMCOBrowsingTerm", '') != '':
    	coldescgrpcontrolaccessriamcoheadelement.text = "RIAMCO Browsing Term"

    repeatingsubjectfield(coldescgrpriamcoscontrolaccesselement, cldata, 'RIAMCOBrowsingTerm', 'subject', {'altrender':'nodisplay','source':'riamco'})

    #Create the container list.
    dscelement = etree.SubElement(archdescelement, "dsc", {"type":"combined"})
    ctelement = dscelement
    crecordgroupelement = dscelement
    csubgroupelement = dscelement
    cserieselement = dscelement
    csubserieselement = dscelement
    csubsubserieselement = dscelement
    cfileelement = dscelement
    rowindex = 2
    onlySeriesRows = True

    #Set up in the case that there is only series and subseries.
    arrangementnotepforlists = etree.SubElement(colarrangementelement, "p")
    cserieslist = etree.SubElement(arrangementnotepforlists, "list")
    csubserieslist = etree.SubElement(cserieslist, "list")

    for row in csvdata:
        if row.get("recordgroupTitle", '') != '' or row.get("subgroupTitle", '') != '' or row.get("subSeriesTitle", '') != '':
            onlySeriesRows = False
            break

    if onlySeriesRows == True:
        arrangementnoteseriespelement = etree.SubElement(colarrangementelement, "p")
        cserieslist = etree.SubElement(arrangementnoteseriespelement, "list")
        print("Only series rows is True.", file=sys.stderr)

    for row in csvdata:
        #eadtop = etree.Element("{http://www.loc.gov/mods/v3}mods", {attr_qname: "http://www.loc.gov/mods/v3 http://www.loc.gov/mods/v3/mods-3-7.xsd"}, nsmap=ns_map)

        if row.get('Ignore', '') != '':
            continue
        if row.get("recordgroupTitle", '') != "":
            serieselement = etree.SubElement(dscelement, "c", {"id":("c"+str(rowindex)), "level":"recordgrp"})
            crecordgroupelement = serieselement
            csubgroupelement = serieselement
            cserieselement = serieselement
            csubserieselement = serieselement
            cfileelement = serieselement
            ctelement = serieselement

            didelement = etree.SubElement(ctelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("recordgroupTitle", '').split())
            cunittitle = titleelement

            seriesIDelement = etree.SubElement(cdid, "unitid", {"type":"recordgrp"})
            cseriesID = "Record Group " + ' '.join(str(row.get("recordgroupID", '')).split()).replace('.0','')
            seriesIDelement.text = cseriesID

            #MODS: subtitle = etree.SubElement(titleinfo, "subTitle")
            #MODS: subtitle.text = ' '.join(row.get("subTitle", '').split())

            #Add the series to the Arrangement Note.
            arrangementnotepaddition = etree.SubElement(colarrangementelement, "p")
            arrangementnoteemphaddition = etree.SubElement(arrangementnotepaddition, "emph", {"render":"bold"})
            arrangementnoteemphaddition.text = cseriesID + ". " + titleelement.text

            arrangementnotepforlists = etree.SubElement(colarrangementelement, "p")
            cserieslist = etree.SubElement(arrangementnotepforlists, "list")
            csubserieslist = etree.SubElement(cserieslist, "list")
        if row.get("subgroupTitle", '') != "":
            serieselement = etree.SubElement(crecordgroupelement, "c", {"id":("c"+str(rowindex)), "level":"subgrp"})
            csubgroupelement = serieselement
            cserieselement = serieselement
            csubserieselement = serieselement
            cfileelement = serieselement
            ctelement = serieselement

            didelement = etree.SubElement(ctelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("subgroupTitle", '').split())
            cunittitle = titleelement

            seriesIDelement = etree.SubElement(cdid, "unitid", {"type":"subgrp"})
            cseriesID = "Subgroup " + ' '.join(str(row.get("subgroupID", '')).split()).replace('.0','')
            seriesIDelement.text = cseriesID

            #MODS: subtitle = etree.SubElement(titleinfo, "subTitle")
            #MODS: subtitle.text = ' '.join(row.get("subTitle", '').split())

            #Add the series to the Arrangement Note.
            arrangementnotepaddition = etree.SubElement(colarrangementelement, "p")
            arrangementnotepaddition.text = cseriesID + ". " + titleelement.text

            arrangementnotepforlists = etree.SubElement(colarrangementelement, "p")
            cserieslist = etree.SubElement(arrangementnotepforlists, "list")
            csubserieslist = etree.SubElement(cserieslist, "list")
        #Create a top level series element if the series cell is not blank.
        if row.get("seriesTitle", '') != "":
            serieselement = etree.SubElement(csubgroupelement, "c", {"id":("c"+str(rowindex)), "level":"series"})
            cserieselement = serieselement
            csubserieselement = serieselement
            cfileelement = serieselement
            ctelement = serieselement

            didelement = etree.SubElement(ctelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("seriesTitle", '').split())
            cunittitle = titleelement

            seriesIDelement = etree.SubElement(cdid, "unitid", {"type":"series"})
            cseriesID = "Series " + ' '.join(str(row.get("seriesID", '')).split()).replace('.0','')
            seriesIDelement.text = cseriesID

            #MODS: subtitle = etree.SubElement(titleinfo, "subTitle")
            #MODS: subtitle.text = ' '.join(row.get("subTitle", '').split())

            #Add the series to the Arrangement Note.
            #arrangementnotepaddition = etree.SubElement(colarrangementelement, "p")
            #arrangementnotepaddition.text = cseriesID + ". " + titleelement.text
            #csubserieslist = etree.SubElement(colarrangementelement, "list")

            seriesarrangementelement = etree.SubElement(cserieslist, "item")
            seriesarrangementelement.text = cseriesID + ". " + titleelement.text

            csubserieslist = etree.SubElement(seriesarrangementelement, "list")

        #Create a top level subseries element if the subseries cell is not blank.
        elif row.get("subSeriesTitle", '') != "":
            subserieselement = etree.SubElement(cserieselement, "c", {"id":("c"+str(rowindex)), "level":"subseries"})
            csubserieselement = subserieselement
            cfileelement = subserieselement
            ctelement = subserieselement

            didelement = etree.SubElement(subserieselement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("subSeriesTitle", '').split())
            cunittitle = titleelement

            subseriesIDelement = etree.SubElement(cdid, "unitid", {"type":"subseries"})
            subseriesIDelement.text = cseriesID + ". Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','')

            #MODS: subtitle = etree.SubElement(titleinfo, "subTitle")
            #MODS: subtitle.text = ' '.join(row.get("subTitle", '').split())

            #arrangementnotepaddition = etree.SubElement(csubserieslist, "item")
            #arrangementnotepaddition.text =  "Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','') + ". " + titleelement.text

            subseriesarrangementelement = etree.SubElement(csubserieslist, "item")
            subseriesarrangementelement.text = "Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','') + ". " + titleelement.text
        #Create a top level file element if the title cell is not blank.
        elif row.get("fileTitle", '') != "":
            fileement = etree.SubElement(csubserieselement, "c", {"id":("c"+str(rowindex)), "level":"file"})
            cfileelement = fileement
            ctelement = fileement

            didelement = etree.SubElement(fileement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("fileTitle", '').split())
            cunittitle = titleelement

            #subseriesIDelement = etree.SubElement(cdid, "unitid", {"type":"subseries"})
            #subseriesIDelement.text = cseriesID + ". Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','')

            #MODS: subtitle = etree.SubElement(titleinfo, "subTitle")
            #MODS: subtitle.text = ' '.join(row.get("subTitle", '').split())

            #arrangementnotepaddition = etree.SubElement(csubserieslist, "item")
            #arrangementnotepaddition.text =  "Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','') + ". " + titleelement.text

        #Create a top level item element if the title cell is not blank.
        elif row.get("itemTitle", '') != "":
            itemelement = etree.SubElement(cfileelement, "c", {"id":("c"+str(rowindex)), "level":"item"})
            ctelement = itemelement

            didelement = etree.SubElement(itemelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("itemTitle", '').split())
            cunittitle = titleelement
            #Needs attention: Item ID.
            #itemIDelement = etree.SubElement(didelement, "unitid", {"type":"subseries"})
            #itemIDelement.text = cseriesID + ". Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','')

        #container
        shelfLocatorstring = ""
        barcodestring = ''

        #if row.get("barcode", '') != '':
        #    floatbarcode = float(row.get("barcode", ''))

        if row.get("barcode", '') != '':
            barcodestring = ' [' + xmltext(row.get("barcode", '')).rstrip('.0') + ']'

        if row.get("shelfLocator1", '') != "":
            # , "label": ' '.join(row.get("shelfLocator1", '').split()) , "label": xmltext(shelfLocator1)
            shelflocator1attributes = {"type":' '.join(row.get("shelfLocator1", '').split()).lower().replace(' ', '_'), "label": xmltext(row.get("shelfLocator1", '').title()) + barcodestring}
            shelflocator1element = etree.SubElement(cdid, "container", shelflocator1attributes)
            shelflocator1element.text = ' '.join(str(row.get("shelfLocator1ID", '')).split()).replace('.0','')
        if row.get("shelfLocator2", '') != "":
            shelflocator2attributes = {"type":' '.join(row.get("shelfLocator2", '').split()).lower().replace(' ', '_'), "label": xmltext(row.get("shelfLocator2", '').title())}
            shelflocator2element = etree.SubElement(cdid, "container", shelflocator2attributes)
            shelflocator2element.text = ' '.join(str(row.get("shelfLocator2ID", '')).split()).replace('.0','')
        if row.get("shelfLocator3", '') != "":
            shelflocator3attributes = {"type":' '.join(row.get("shelfLocator3", '').split()).lower().replace(' ', '_'), "label": xmltext(row.get("shelfLocator3", '').title())}
            shelflocator3element = etree.SubElement(cdid, "container", shelflocator3attributes)
            shelflocator3element.text = ' '.join(str(row.get("shelfLocator3ID", '')).split()).replace('.0','')

        #dates
        #Test for a YYYY - YYYY and remove dates if so.
        match = re.search(u"(\d{4}\s-\s\d{4})", row.get("dateText", ''))

        inclusivedatetext= xmltext(row.get("dateText", '')).replace('.0','') #' '.join(row.get("dateText", '').split()).replace('.0','')
        if match:
            inclusivedatetext = inclusivedatetext.replace(' ','')
        inclusivedatestart = xmltext(row.get("dateStart", '')).replace('.0','') #' '.join(row.get("dateStart", '').split()).replace('.0','')
        inclusivedateend = xmltext(row.get("dateEnd", '')).replace('.0','') #' '.join(row.get("dateEnd", '').split()).replace('.0','')

        bulkdatestart = xmltext(row.get("dateBulkStart", '')).replace('.0','') # ' '.join(row.get("dateBulkStart", '').split()).replace('.0','')
        bulkdateend = xmltext(row.get("dateBulkEnd", '')).replace('.0','') #' '.join(row.get("dateBulkEnd", '').split()).replace('.0','')

        unitdateinclusiveattributes = {"type":"inclusive"}
        if inclusivedatestart != '' and inclusivedateend != '':
        	unitdateinclusiveattributes["normal"] = inclusivedatestart +"/"+inclusivedateend
        if row.get("dateQualifier", '') != "":
            unitdateinclusiveattributes["certainty"] = xmltext(row.get("dateQualifier", ''))
        unitdatebulkattributes = {"type":"bulk","normal":(bulkdatestart +"/"+bulkdateend)}

        inclusivedateelement = etree.SubElement(cdid, "unitdate", unitdateinclusiveattributes)
        if inclusivedatetext == '' and (inclusivedatestart.count != '' or inclusivedateend != ''):
            if inclusivedatestart != inclusivedateend:
                inclusivedatetext = inclusivedatestart + ' - ' + inclusivedateend
                inclusivedatetext = inclusivedatetext.lstrip(' - ').rstrip(' - ')
            else:
                inclusivedatetext = inclusivedatestart
        inclusivedateelement.text = inclusivedatetext

        if bulkdatestart != "" or bulkdateend != "":
            bulkdateelement = etree.SubElement(cdid, "unitdate", unitdatebulkattributes)
            bulkdateelement.text = "(bulk " + bulkdatestart + "-" + bulkdateend + ")"

        #physicalDescription
        physicalDescriptionelement = etree.SubElement(cdid, "physdesc")

        #extent and genre
        extentquantityphysdescelement = etree.SubElement(cdid, "physdesc", {"altrender":"whole"})
        extentQuantityelement = etree.SubElement(extentquantityphysdescelement, "extent", {"altrender":"materialtype spaceoccupied"})
        extentQuantityelement.text = ' '.join(row.get("extentQuantity", '').split())
        containerSummaryelement = etree.SubElement(extentquantityphysdescelement, "extent",{"altrender":"carrier"})
        containerSummaryelement.text = xmltext(row.get("containerSummary",''))

        extentsizephysdescelement = etree.SubElement(cdid, "physdesc")
        extentSizeelement = etree.SubElement(extentsizephysdescelement, "dimensions")
        extentSizeelement.text = ' '.join(row.get("extentSize", '').split())

        extentSpeedelement = etree.SubElement(extentsizephysdescelement, "dimensions")
        extentSpeedelement.text = ' '.join(row.get("extentSpeed", '').split())

        genreformphyscdescelement = etree.SubElement(cdid, "physdesc")
        repeatingsubjectfield(genreformphyscdescelement, row, 'genreAAT', 'genreform',{"source":"aat"})
        repeatingsubjectfield(genreformphyscdescelement, row, 'genreLCSH', 'genreform',{"source":"lcsh"})
        repeatingsubjectfield(genreformphyscdescelement, row, 'genreLocal', 'genreform',{"source":"local"})
        repeatingsubjectfield(genreformphyscdescelement, row, 'genreRBGENR', 'genreform',{"source":"rbgenr"})

        #materialspec
        formelement = etree.SubElement(cdid, "materialspec")
        formelement.text = ' '.join(row.get("form", '').split())

        #language
        langmaterialelement = etree.SubElement(cdid, "langmaterial")
        langmaterialelementsplitchar = getSplitCharacter(row.get("language", ''))

        for language in row.get("language", '').split(langmaterialelementsplitchar):
            if language == "":
                continue
            if len(xmltext(language)) < 4:
                language = langcodeopp.get(language, '')
            if xmltext(language) in langcode:
                langusagelangelement = etree.SubElement(langmaterialelement, "language", {"langcode":langcode[xmltext(language)], "scriptcode":scriptcode[xmltext(language)]})
                langusagelangelement.text =  xmltext(language)

                if scriptcode == "N/A":
                    langissue = True
            else:
                langusagelangelement = etree.SubElement(langmaterialelement, "language", {"langcode":"***", "scriptcode":"***"})
                langusagelangelement.text =  xmltext(language)

                langissue = True

        #Create origination and controlaccess element.

        originationelement = etree.SubElement(cdid, "origination")

        #typeOfResource
        #Needs attention: Instance Type/typeOfResource.
        #typeofresource = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}typeOfResource")
        #typeofresource.text = ' '.join(row.get("typeOfResource", '').split())

        #note
        notescopeelement = etree.SubElement(ctelement, "scopecontent")
        scopelines = row.get("noteScope", '').splitlines()
        for line in scopelines:
            pelement = etree.SubElement(notescopeelement, "p")
            pelement.text = ' '.join(line.split())

        notebioghistelement = etree.SubElement(ctelement, "bioghist")
        bioghistlines = row.get("noteHistorical", '').splitlines()
        for line in bioghistlines:
            pelement = etree.SubElement(notebioghistelement, "p")
            pelement.text = ' '.join(line.split())

        acqinfoelement = etree.SubElement(ctelement, "acqinfo")
        noteaccessionpelement = etree.SubElement(acqinfoelement, "p")
        noteaccessionelement = etree.SubElement(noteaccessionpelement, "num", {"type":"accession"})
        noteaccessionelement.text = ' '.join(row.get("noteAccession", '').split())

        #useAndReproduction
        #useAndReproductionelement = etree.SubElement(ctelement, "userestrict")
        #useAndReproductionelementlines = row.get("useAndReproduction", '').splitlines()
        #for line in useAndReproductionelementlines:
        #    pelement = etree.SubElement(useAndReproductionelement, "p")
        #    pelement.text = ' '.join(line.split())

        #controlaccess and name fields
        ccontrolaccess = etree.SubElement(ctelement, "controlaccess")

        repeatingNameField(originationelement, "persname", row.get("namePersonCreatorLC", ''), "creator", "naf")
        repeatingNameField(originationelement, "persname", row.get("namePersonCreatorLocal", ''), "creator", "local")
        repeatingNameField(originationelement, "persname", row.get("namePersonCreatorFAST", ''), "creator", "fast")

        repeatingNameField(originationelement, "corpname", row.get("nameCorpCreatorLC", ''), "creator", "naf")
        repeatingNameField(originationelement, "corpname", row.get("nameCorpCreatorLocal", ''), "creator", "local")
        repeatingNameField(originationelement, "corpname", row.get("nameCorpCreatorFAST", ''), "creator", "fast")

        repeatingNameField(ccontrolaccess, "persname", row.get("namePersonOtherLC", ''), "", "naf")
        repeatingNameField(ccontrolaccess, "persname", row.get("namePersonOtherLocal", ''), "", "local")
        repeatingNameField(ccontrolaccess, "persname", row.get("namePersonOtherFAST", ''), "", "fast")

        notegeneralelement = etree.SubElement(ctelement, "odd")
        scopelines = row.get("noteGeneral", '').splitlines()
        for line in scopelines:
            pelement = etree.SubElement(notegeneralelement, "p")
            pelement.text = ' '.join(line.split())

        notealtformelement = etree.SubElement(ctelement, "altformavail")
        altformlines = row.get("locationCopies", '').splitlines()
        for line in altformlines:
            pelement = etree.SubElement(notealtformelement, "p")
            pelement.text = ' '.join(line.split())

        #noteAccessionelement = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}note", {"type":"acquisition", "displayLabel":"Immediate form of acquisition"})
        #noteAccessionelement.text = ' '.join(row.get("noteAccession", '').split())

        #noteHistoricalClassYearelement = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}note", {"type":"biographical/historical", "displayLabel":"Class year"})
        #noteHistoricalClassYearelement.text = ' '.join(row.get("noteHistoricalClassYear", '').split())

        #notePreferredCitation = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}note", {"type":"preferredCitation"})
        #notePreferredCitationstring = ' '.join(row.get("title", '').split()).rstrip('.')
        #if row.get("collection", '') != "":
        #    notePreferredCitationstring = notePreferredCitationstring + ", " + ' '.join(row.get("collection", '').split())
        #if row.get("callNumber", '') != "":
        #    notePreferredCitationstring = notePreferredCitationstring + ", " + ' '.join(row.get("callNumber", '').split())
        #notePreferredCitation.text = notePreferredCitationstring + ', Brown University Library'

        #originInfo

        #MODS: publisherelement = etree.SubElement(originInfoelement, "{http://www.loc.gov/mods/v3}publisher")
        #MODS: publisherelement.text = ' '.join(row.get("publisher", '').split())

        #dateQualifierAttribute = {}

        #if row.get("dateQualifier", '') != "":
        #    dateQualifierAttribute = {"qualifier": row.get("dateQualifier", '')}


        #geogname
        geognameelement = etree.SubElement(cunittitle, "geogname")
        geognameelement.text = ' '.join(row.get("place", '').split())

        #rightsStatementelement = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}accessCondition", {"type":"rightsStatement","{http://www.w3.org/1999/xlink}href":row.get("rightsStatementURI", '')})
        #rightsStatementelement.text = ' '.join(row.get("rightsStatementText", '').split())

        #restrictionOnAccesselement = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}accessCondition", {"type":"restrictionOnAccess"})
        #restrictionOnAccesselement.text = "Collection is open for research."

        #subject
        repeatingsubjectfield(ccontrolaccess, row, 'subjectNamesLC', 'persname',{"source":"naf"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectNamesLocal', 'persname',{"source":"local"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectNamesFAST', 'persname',{"source":"fast"})

        repeatingsubjectfield(ccontrolaccess, row, 'subjectCorpLC', 'corpname',{"source":"naf"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectCorpLocal', 'corpname',{"source":"local"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectCorpFAST', 'corpname',{"source":"fast"})

        repeatingsubjectfield(ccontrolaccess, row, 'subjectTopicsLC', 'subject',{"source":"lcsh"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectTopicsLocal', 'subject',{"source":"local"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectTopicsFAST', 'subject',{"source":"fast"})

        repeatingsubjectfield(ccontrolaccess, row, 'subjectGeoLC', 'geogname',{"source":"lcsh"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectGeoFAST', 'geogname',{"source":"fast"})

        repeatingsubjectfield(ccontrolaccess, row, 'subjectTemporalLC', 'subject',{"source":"lcsh"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectTemporalFAST', 'subject',{"source":"fast"})

        repeatingsubjectfield(ccontrolaccess, row, 'subjectTitleLC', 'title',{"source":"lcsh"})
        repeatingsubjectfield(ccontrolaccess, row, 'subjectTitleFAST', 'title',{"source":"fast"})

        #for index, temporal in enumerate(row.get("subjectTemporalLC", '').split('|')):
        #    subjectelement = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}subject", {"authority":"local"})
        #    if temporal == "":
        #        continue
        #    temporalelement = etree.SubElement(subjectelement, "{http://www.loc.gov/mods/v3}temporal")
        #    temporalelement.text = ' '.join(temporal.split())

        ###

        #hostlocationelement = etree.SubElement(relatedItemelement, "{http://www.loc.gov/mods/v3}location")

        #hostphysicalLocationelement = etree.SubElement(hostlocationelement, "{http://www.loc.gov/mods/v3}physicalLocation")
        #hostphysicalLocationelement.text = ' '.join(row.get("repository", '').split())

        #hosturlelement = etree.SubElement(hostlocationelement, "{http://www.loc.gov/mods/v3}url")
        #hosturlelement.text = ' '.join(row.get("findingAid", '').split())

        #hostholdingSimpleelement = etree.SubElement(hostlocationelement, "{http://www.loc.gov/mods/v3}holdingSimple")
        #hostcopyInformationelement = etree.SubElement(hostholdingSimpleelement, "{http://www.loc.gov/mods/v3}copyInformation")
        #hostshelfLocatorelement = etree.SubElement(hostcopyInformationelement, "{http://www.loc.gov/mods/v3}shelfLocator")


      #  if row.get("shelfLocator2", '') != "":
     #       shelfLocatorstring = shelfLocatorstring + ', ' + ' '.join(row.get("shelfLocator2", '').split()) + ' ' + ' '.join(str(row.get("shelfLocator2ID", '')).split()).replace('.0','')
      #  if row.get("shelfLocator3", '') != "":
      #      shelfLocatorstring = shelfLocatorstring + ', ' + ' '.join(row.get("shelfLocator3", '').split()) + ' ' + ' '.join(str(row.get("shelfLocator3ID", '')).split()).replace('.0','')

     #   hostshelfLocatorelement.text = ' '.join(shelfLocatorstring.split())

        #dao fields
        if row.get("identifierBDR", '') != "":
            daomodselement = etree.SubElement(ctelement, "dao", {"{http://www.w3.org/1999/xlink}actuate":"onRequest","{http://www.w3.org/1999/xlink}show":"embed","{http://www.w3.org/1999/xlink}title": cunittitle.text, "{http://www.w3.org/1999/xlink}role":"MODS_ID","{http://www.w3.org/1999/xlink}href":'bdr'+ xmltext(row.get("identifierBDR", '')).lstrip('bdr').replace(':','')})
            daomodsdescelement = etree.SubElement(daomodselement, "daodesc")
            daomodspelement = etree.SubElement(daomodsdescelement, "p")
            daomodspelement.text = cunittitle.text

            daobdrelement = etree.SubElement(ctelement, "dao", {"{http://www.w3.org/1999/xlink}actuate":"onRequest","{http://www.w3.org/1999/xlink}show":"embed","{http://www.w3.org/1999/xlink}title": cunittitle.text, "{http://www.w3.org/1999/xlink}role":"BDR_PID","{http://www.w3.org/1999/xlink}href":'bdr:'+ ' '.join(row.get("identifierBDR", '').split()).lstrip('bdr').replace(':','')})
            daobdrdescelement = etree.SubElement(daobdrelement, "daodesc")
            daobdrpelement = etree.SubElement(daobdrdescelement, "p")
            daobdrpelement.text = cunittitle.text

        if row.get("identifierNormalized", '') != "":
            daomodselement = etree.SubElement(ctelement, "dao", {"{http://www.w3.org/1999/xlink}actuate":"onRequest","{http://www.w3.org/1999/xlink}show":"embed","{http://www.w3.org/1999/xlink}title": cunittitle.text, "{http://www.w3.org/1999/xlink}role":"NORMALIZEDFILE_ID","{http://www.w3.org/1999/xlink}href": xmltext(row.get("identifierNormalized", '')).lstrip('bdr')})
            daomodsdescelement = etree.SubElement(daomodselement, "daodesc")
            daomodspelement = etree.SubElement(daomodsdescelement, "p")
            daomodspelement.text = cunittitle.text

        if row.get("identifierWebArchive", '') != "":
            daomodselement = etree.SubElement(ctelement, "dao", {"{http://www.w3.org/1999/xlink}actuate":"onRequest","{http://www.w3.org/1999/xlink}show":"embed","{http://www.w3.org/1999/xlink}title": cunittitle.text, "{http://www.w3.org/1999/xlink}role":"WEBARCHIVEURL","{http://www.w3.org/1999/xlink}href":xmltext(row.get("identifierWebArchive", '')).lstrip('bdr')})
            daomodsdescelement = etree.SubElement(daomodselement, "daodesc")
            daomodspelement = etree.SubElement(daomodsdescelement, "p")
            daomodspelement.text = cunittitle.text


        #MODS: lastnote
        #digitalObjectMadeelement = etree.SubElement(eadtop, "{http://www.loc.gov/mods/v3}note", {"displayLabel":"Digital object made available by"})
        #digitalObjectMadeelement.text = "Brown University Library, John Hay Library, University Archives and Manuscripts, Box A, Brown University, Providence, RI, 02912, U.S.A., (http://library.brown.edu/)"

        rowindex = rowindex + 1

    #Create the collection-level data.
    eadidattributes = {"countrycode":"US", "mainagencycode":"US-" + xmltext(cldata.get("MARCRepositoryCode", '')), "identifier":xmltext(cldata.get("callNumber", '')).lower()+'.xml'}
    eadidelement = etree.SubElement(eadheaderelement, "eadid", eadidattributes)
    eadidelement.text = "US-"+xmltext(cldata.get("MARCRepositoryCode", ''))+"-"+xmltext(cldata.get("callNumber", '')).lower()

    filedescelement = etree.SubElement(eadheaderelement, "filedesc")

    #titlestmt
    titlestmtelement = etree.SubElement(filedescelement, "titlestmt")
    titleproperelement = etree.SubElement(titlestmtelement, "titleproper")
    titleproperelement.text = "Guide to the " + xmltext(cldata.get("title", ''))

    inclusivedateattributes = {"type":"inclusive", "era":"ce","calendar":"gregorian","normal":xmltext(cldata.get("inclusiveDates", '')).replace('-','/').replace(' ','')}
    inclusivedateelement = etree.SubElement(titleproperelement, "date", inclusivedateattributes)
    inclusivedateelement.text = xmltext(cldata.get("inclusiveDates", '').replace(' ',''))

    bulkdateattributes = {"type":"bulk", "era":"ce","calendar":"gregorian","normal":xmltext(cldata.get("bulkDates", '')).replace('-','/').replace(' ','')}
    bulkdateelement = etree.SubElement(titleproperelement, "date", bulkdateattributes)
    bulkdateelement.text = "(bulk "+ xmltext(cldata.get("bulkDates", '')).replace(' ','') + ")"

    authorelement = etree.SubElement(titlestmtelement, "author").text = "Finding aid prepared by " + xmltext(cldata.get("author", ''))

    sponsorelement = etree.SubElement(titlestmtelement, "sponsor").text = xmltext(cldata.get("sponsor", ''))

    #publicationstmt
    publicationstmtelement = etree.SubElement(filedescelement, "publicationstmt")

    publisherelement = etree.SubElement(publicationstmtelement, "publisher")
    publisherelement.text = xmltext(cldata.get("publisher", ''))

    pubaddresselement = etree.SubElement(publicationstmtelement, "address")

    pubaddresslines = cldata.get("address", '').splitlines()
    for line in pubaddresslines:
        addresselement = etree.SubElement(pubaddresselement, "addressline")
        addresselement.text = xmltext(line)

    creationdateelement = etree.SubElement(publicationstmtelement, "date", {"era":"ce","calendar":"gregorian", "normal":xmltext(cldata.get("creationDate", ''))[:4], "type":"publication"})
    creationdateelement.text = xmltext(cldata.get("creationDate", '').replace('.0',''))

    #profiledesc
    profiledescelement = etree.SubElement(eadheaderelement, "profiledesc")

    creationelement = etree.SubElement(profiledescelement, "creation")
    creationelement.text = "This finding aid was produced using the RIAMCO EAD spreadsheet, "
    creationdatecreationelement = etree.SubElement(creationelement, "date", {"era":"ce","calendar":"gregorian", "normal":xmltext(cldata.get("creationDate", ''))[:4], "type":"publication"})
    creationdatecreationelement.text = xmltext(cldata.get("creationDate", '').replace('.0',''))

    #langusage
    langusageelement = etree.SubElement(profiledescelement, "langusage")

    if xmltext(cldata.get("findingAidLanguage", '')) in langcode:
        langusagelangelement = etree.SubElement(langusageelement, "language", {"langcode":langcode[xmltext(cldata.get("findingAidLanguage", ''))], "scriptcode":scriptcode[xmltext(cldata.get("findingAidLanguage", ''))]})
        langusagelangelement.text = xmltext(cldata.get("findingAidLanguage", ''))

        if scriptcode == "N/A":
            langissue = True
    else:
        langusagelangelement = etree.SubElement(langusageelement, "language", {"langcode":"***", "scriptcode":"***"})
        langusagelangelement.text = xmltext(cldata.get("findingAidLanguage", ''))

        langissue = True

    descruleselement = etree.SubElement(profiledescelement, "descrules")
    descruleselement.text = "Finding aid based on Describing Archives: A Content Standard (DACS)"



    # start cleanup
    # remove any element tails
    for element in eadtop.iter():
        element.tail = None

    # remove any line breaks or tabs in element text
        if element.text:
            if '\n' in element.text:
                element.text = element.text.replace('\n', '')
            if '\t' in element.text:
                element.text = element.text.replace('\t', '')

    # remove any remaining whitespace
    parser = etree.XMLParser(remove_blank_text=True, remove_comments=True, recover=True)
    treestring = etree.tostring(eadtop)
    clean = etree.XML(treestring, parser)

    # remove recursively empty nodes
    # found here: https://stackoverflow.com/questions/12694091/python-lxml-how-to-remove-empty-repeated-tags
    def recursively_empty(e):
       if e.text:
           return False
       return all((recursively_empty(c) for c in e.iterchildren()))

    context = etree.iterwalk(clean)
    for action, elem in context:
        parent = elem.getparent()
        if recursively_empty(elem):
            parent.remove(elem)

    # remove nodes with blank attribute
    for element in clean.xpath(".//*[@*='']"):
        element.getparent().remove(element)

    # remove nodes with attribute "null"
    for element in clean.xpath(".//*[@*='null']"):
        element.getparent().remove(element)

    # finished cleanup
    # write out to intermediate file
    #with open(os.path.dirname(os.path.abspath(__file__)) + '/cache/clean.xml', 'wb') as f:
    #    f.write(etree.tostring(clean))
    #print "XML is now clean"

    completestring = etree.tostring(clean, pretty_print = True, encoding="unicode")
    completestring = completestring.replace('&lt;','<')
    completestring = completestring.replace('&gt;','>')

    finalfilename = eadidelement.text + '.xml'

    with open(CACHEDIR + id + "/" + finalfilename, 'w+') as f:
        f.write("<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n")
        f.write(completestring)
        f.close()


    returndict = {}

    returndict["filename"] = finalfilename
    returndict["error"] = False
    returndict["allrecords"] = completestring

    with open(CACHEDIR + id + "/" + finalfilename, 'rb') as f:
        return(f.read(), returndict)

    if langissue == True:
        #print("")
        print("*Language Field Error*\n", file=sys.stderr)
        print("There were one or more issues with language fields in your spreadsheet. Please check your spelling in all language fields. You may also manually correct your XML file, consult the SupportedLanguages.xlsx file in the data folder for supported languages, and/or adjust the SupportedLanguages.xlsx spreadsheet to suit your project.\n", file=sys.stderr)

        print('\n\n', file=sys.stderr)

    #errorfile.close()
    print("***Operation Complete***\n", file=sys.stderr)
    #rint("Your EAD file was written to folder " + output_path + ".\n", file=sys.stderr)

    print('\n\n', file=sys.stderr)

    print( u"   \u2606 \u2606 \u2606", file=sys.stderr)
    print('\n\n\n', file=sys.stderr)

