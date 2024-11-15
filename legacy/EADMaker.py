import os
from lxml import etree
from openpyxl import load_workbook
import re
import xlrd
from copy import copy
import sys
from tools.stringbools import hasLetters, hasYear
from legacy.eadconfig import requiredcolumns

xlrd.xlsx.ensure_elementtree_imported(False, None)
xlrd.xlsx.Element_has_iter = True

langcode = {}
langcodeopp = {}
scriptcode = {}
langissue = False
CACHEDIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache") + "/"
#CACHEDIR = os.path.dirname(os.path.abspath(__file__)) + "/"
HOMEDIR = os.path.dirname(os.path.abspath(__file__)) + "/"
#HOMEDIR = os.path.dirname(os.path.abspath(__file__)) + "/"

def getSplitCharacter(input_str):
    if ";" in input_str:
        return(";")
    else:
        return("|")

def multilinefield(refdict, parentelement, originalfieldname, eadfieldname):
    newelement = etree.SubElement(parentelement, eadfieldname)
    lines = refdict.get(originalfieldname, '').splitlines()
    for line in lines:
        pelement = etree.SubElement(newelement, "p")
        pelement.text = ' '.join(line.split())

def repeatingsubjectfield(parentelement, refdict, originalfieldname, eadfieldname, eadattributes):
    splitcharacter = ";"

    for namesindex, addedentry in enumerate(refdict.get(originalfieldname, '').split(splitcharacter)):

        customAttributes = eadattributes.copy()

        #Extract URI

        uri = re.findall("(?P<url>https?://[^\s]+)", addedentry)

        #If there's a URI
        if len(uri) > 0:
            #Remove it from the addedentry
            addedentry = addedentry.replace(uri[0],"")
            #Add it as a valueURI attribute
            customAttributes["authfilenumber"] = normalize_whitespace(uri[0])

        namecontrolaccesselement = etree.SubElement(parentelement, eadfieldname, customAttributes)
        namecontrolaccesselement.text = ' '.join(addedentry.replace("|d", "").replace("|e", "").split())

def repeatingNameField(parentElement, elementName, rowString, assignedRole, source):
    for name in rowString.split(';'):
        currentname = ""
        currentrole = ""
        attributes = {}

        #Extract URI
        uri = re.findall("(?P<url>https?://[^\s]+)", name)
        if uri:
            #Remove it from the addedentry
            name = name.replace(uri[0],"")
            #Add it as a valueURI attribute
            attributes["authfilenumber"] = normalize_whitespace(uri[0])

        for index, namefield in enumerate(name.split(',')):
            namefieldrevised = normalize_whitespace(namefield)

            if index == 0:
                currentname = currentname + namefieldrevised + ", "
            elif hasYear(namefieldrevised):
                currentname = currentname + namefieldrevised
            elif namefieldrevised.islower():
                currentrole = namefieldrevised
            elif hasLetters(namefieldrevised) is not None:
                currentname = currentname + namefieldrevised + ", "

        if currentrole:
            attributes['role'] = currentrole
        elif assignedRole:
            attributes['role'] = assignedRole

        attributes['source'] = source

        nameelement = etree.SubElement(parentElement, elementName, attributes)
        nameelement.text = normalize_whitespace(currentname).rstrip(',').lstrip(',')

def normalize_whitespace(text, replace_with=' '):
    return(replace_with.join(str(text).split()))

def copyworkbook(path1, path2):

    wb1 = load_workbook(filename=path1)
    ws1 = wb1.worksheets[0]

    wb2 = load_workbook(filename=path2)
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            copycell = ws2[cell.coordinate]
            copycell.value = cell.value
            if cell.has_style:
                copycell.font = copy(cell.font)
                copycell.border = copy(cell.border)
                copycell.fill = copy(cell.fill)
                copycell.number_format = copy(cell.number_format)
                copycell.protection = copy(cell.protection)
                copycell.alignment = copy(cell.alignment)

    wb2.save(path2)

def XLSDictReader(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        rowarray = []

        for row in range(1, sheet.nrows):
            rowdictionary = {}
            for column in range(sheet.ncols):
                #If the value is a number, turn it into a string.
                newvalue = ''
                if sheet.cell(row,column).ctype > 1:
                    newvalue = str(sheet.cell_value(row,column)).replace('|d', '').replace('|e', '').replace('|',';')
                    #print(sheet.cell_value(row,column))
                else:
                    newvalue = sheet.cell_value(row,column).replace('|d', '').replace('|e', '').replace('|',';')

                #If the column is repeating, serialize the row values.
                if rowdictionary.get(sheet.cell_value(0,column), '') != '':
                    rowdictionary[sheet.cell_value(0,column)] = rowdictionary[sheet.cell_value(0,column)] + ';' + newvalue
                else:
                    rowdictionary[sheet.cell_value(0,column)] = newvalue
            rowarray.append(rowdictionary)
        return(rowarray)

def XLSDictReaderVertical(file, sheetname):
        book    = xlrd.open_workbook(file)
        sheet   = book.sheet_by_name(sheetname)

        verticaldictionary = {}
        datacolumn = 3

        for column in range(0,sheet.ncols):
            #print(sheet.cell_value(0, column))
            if "Data Entry" in sheet.cell_value(0, column):
                datacolumn = column

        for row in range(sheet.nrows):
            key = sheet.cell_value(row, 0)
            value = ""
            if sheet.cell(row,datacolumn).ctype > 1:
                    value = str(sheet.cell_value(row,datacolumn)).replace('|d', '').replace('|e', '').replace('|',';')
            else:
                    value = sheet.cell_value(row,datacolumn).replace('|d', '').replace('|e', '').replace('|',';')
            if verticaldictionary.get(key, '') == '':
                verticaldictionary[key] = value
            else:
                verticaldictionary[key] = verticaldictionary[key] + ';' + value
        return(verticaldictionary)

def GetXLSDictReaderCodes(file, sheetname):
    book    = xlrd.open_workbook(file)
    sheet   = book.sheet_by_name(sheetname)

    langcode = {}
    langcodeopp = {}
    scriptcode = {}

    for row in range(sheet.nrows):
        for codename, tuple in {"langcode":(0,1),"langcodeopp":(1,0),"scriptcode":(0,2)}.items():
            key = sheet.cell_value(row, tuple[0])
            value = sheet.cell_value(row, tuple[1])
            if codename == "langcode":
                langcode[key] = value
            elif codename == "langcodeopp":
                langcodeopp[key] = value
            else:
                scriptcode[key] = value

    return(langcode, langcodeopp, scriptcode)

def getSheetNames(chosenfile):
    excel = xlrd.open_workbook(chosenfile)
    sheetnames = excel.sheet_names()
    return(sheetnames)

print("._. EAD Maker ._.", file=sys.stderr)

def make_dao_element(ctelement, cunittitle, role, href):
    default_dao_attributes = {
        "{http://www.w3.org/1999/xlink}actuate":"onRequest",
        "{http://www.w3.org/1999/xlink}show":"embed",
        "{http://www.w3.org/1999/xlink}title": cunittitle.text
    }
    dao_dict = {
        **default_dao_attributes,
        "{http://www.w3.org/1999/xlink}role": role,
        "{http://www.w3.org/1999/xlink}href": href
    }
    daomodselement = etree.SubElement(ctelement, "dao", dao_dict)
    daomodsdescelement = etree.SubElement(daomodselement, "daodesc")
    daomodspelement = etree.SubElement(daomodsdescelement, "p")
    daomodspelement.text = cunittitle.text

def processExceltoEAD(chosenfile, chosensheet, id):
    #Create a cache directory for the current EAD file.
    if not os.path.exists(CACHEDIR + id):
            os.mkdir(CACHEDIR + id)

    #Get all languages codes and script codes.
    langcode, langcodeopp, scriptcode = GetXLSDictReaderCodes(HOMEDIR + "SupportedLanguages.xlsx","languages xlsx")

    csvdata = {}
    cldata = {}
    langissue = False

    excel = xlrd.open_workbook(chosenfile)
    sheetnames = excel.sheet_names()
    selectedsheet = excel.sheet_by_name(chosensheet)
    columnsinsheet = [str(cell.value) for cell in selectedsheet.row(0)]

    # Check for missing columns.
    missingcolumns = [ col for col in requiredcolumns if col not in columnsinsheet]
    if missingcolumns:
        print(
            "*Missing Required Columns Detected*\n",
            "The required columns below are missing from your spreadsheet. The script will continue without them.\n\n",
            file=sys.stderr
        )

        for column in missingcolumns:
            print(f"   {column}\n", file=sys.stderr)

        print('\n\n', file=sys.stderr)

    csvdata = XLSDictReader(chosenfile, chosensheet)

    if "Collection-Level Data" not in sheetnames:
        copyworkbook(HOMEDIR + "Collection-Level Data.xlsx", chosenfile)
        excel = xlrd.open_workbook(chosenfile)

        print("*Collection-Level Data Missing*\n", file=sys.stderr)
        print("Collection-level data is missing from your spreadsheet. A sheet titled Collection-Level Data has been automatically added. Enter data in this sheet to add collection-level data to your EAD file.\n", file=sys.stderr)


        print('\n\n', file=sys.stderr)

    cldata = XLSDictReaderVertical(chosenfile, "Collection-Level Data")
    chosenfile = chosensheet

    #Set up namespaces and attributes for XML.
    attr_qname = etree.QName("http://www.w3.org/2001/XMLSchema-instance", "schemaLocation")
    ns_map = {"ead":"urn:isbn:1-931666-22-9", "ns2" : "http://www.w3.org/1999/xlink", "xsi" : "http://www.w3.org/2001/XMLSchema-instance"}
    ns_map2 = {"ns2":"http://www.w3.org/1999/xlink"}
    #Create top elements for EAD.
    eadtop = etree.Element("ead", {attr_qname: "urn:isbn:1-931666-22-9 http://www.loc.gov/ead/ead.xsd", "audience":"external","relatedencoding":"MARC21", "xmlns":"urn:isbn:1-931666-22-9"}, nsmap=ns_map)
    eadheaderelement = etree.SubElement(eadtop, "eadheader", {"audience":"external","countryencoding":"iso3166-1","dateencoding":"iso8601","scriptencoding":"iso15924", "relatedencoding":"MARC21", "repositoryencoding":"iso15511","langencoding":"iso639-2b"})

    #Create archival description elements.
    archdescelement = etree.SubElement(eadtop, "archdesc", {"level":"collection", "type":"inventory"})
    coldidelement = etree.SubElement(archdescelement, "did")

    #Create dsc element and the ctelement variable, which will hold the current series or subseries the script is adding files/items to.
    cseriesID = ""
    ccontrolaccess = etree.Element("ignore")
    cdid = etree.Element("ignoreagain")
    cunittitle = etree.Element("ignoreagainagain")
    cserieslist = etree.Element("continuetoignore")
    csubserieslist = etree.Element("ignoreagainagain")

    #Collection-level description in the archdesc element.
    primaryunittitle = etree.SubElement(coldidelement, "unittitle", {"type":"primary"})
    primaryunittitle.text = normalize_whitespace(cldata.get("title", ''))

    filingunittitle = etree.SubElement(coldidelement, "unittitle", {"type":"filing"})
    filingunittitle.text = normalize_whitespace(cldata.get("filingTitle", ''))

    colunitidelement = etree.SubElement(coldidelement, "unitid", {"countrycode":"US","repositorycode":"US-"+normalize_whitespace(cldata.get("MARCRepositoryCode", '')),"type":"collection"})
    colunitidelement.text = normalize_whitespace(cldata.get("callNumber", ''))

    colrepositoryelement = etree.SubElement(coldidelement, "repository")
    repositorycorpelement = etree.SubElement(colrepositoryelement, "corpname")
    repositorycorpelement.text = normalize_whitespace(cldata.get("repositoryCorporateName", ''))
    repositorysubarea = etree.SubElement(repositorycorpelement, "subarea")
    repositorysubarea.text = normalize_whitespace(cldata.get("repositoryCorporateSubarea", ''))

    coladdresselement = etree.SubElement(colrepositoryelement, "address")
    coladdresslines = cldata.get("repositoryAddress", '').splitlines()
    for line in coladdresslines:
        addresselement = etree.SubElement(coladdresselement, "addressline")
        addresselement.text = normalize_whitespace(line)

    collangmaterial = etree.SubElement(coldidelement, "langmaterial")
    collangmaterialsplitchar = getSplitCharacter(cldata.get("materialLanguage", ''))

    for language in cldata.get("materialLanguage", '').split(collangmaterialsplitchar):
        if language == "":
             continue
        if normalize_whitespace(language) in langcode:
             langusagelangelement = etree.SubElement(collangmaterial, "language", {"langcode":langcode[normalize_whitespace(language)], "scriptcode":scriptcode[normalize_whitespace(language)]})
             langusagelangelement.text =  normalize_whitespace(language)

             if scriptcode == "N/A":
                 langissue = True
        else:
             langusagelangelement = etree.SubElement(collangmaterial, "language", {"langcode":"***", "scriptcode":"***"})
             langusagelangelement.text =  normalize_whitespace(language)

             langissue = True

    colphysdescelement = etree.SubElement(coldidelement, "physdesc")
    colextentelement = etree.SubElement(colphysdescelement, "extent").text = normalize_whitespace(cldata.get("sizeExtent", ''))

    inclusivedateattributes = {"type":"inclusive", "era":"ce","calendar":"gregorian","normal":normalize_whitespace(cldata.get("inclusiveDates", '')).replace('-','/').replace(' ','')}
    inclusivedateelement = etree.SubElement(coldidelement, "unitdate", inclusivedateattributes)
    inclusivedateelement.text = normalize_whitespace(cldata.get("inclusiveDates", '').replace(' ',''))

    bulkdateattributes = {"type":"bulk", "era":"ce","calendar":"gregorian","normal":normalize_whitespace(cldata.get("bulkDates", '')).replace('-','/').replace(' ','')}
    bulkdateelement = etree.SubElement(coldidelement, "unitdate", bulkdateattributes)
    bulkdateelement.text = "(bulk "+ normalize_whitespace(cldata.get("bulkDates", '').replace(' ','')) + ")"

    coloriginationelement = etree.SubElement(coldidelement, "origination", {"label":"creator"})

    colpersoncreatorelement = etree.SubElement(coloriginationelement, "persname", {"role":"creator"})
    colpersoncreatorelement.text = normalize_whitespace(cldata.get("creatorPerson", '')).replace('|d','')

    colcorporatecreatorelement = etree.SubElement(coloriginationelement, "corpname", {"role":"creator"})
    colcorporatecreatorelement.text = normalize_whitespace(cldata.get("creatorCorporate", ''))

    colabstractelement = etree.SubElement(coldidelement, "abstract")
    colabstractelement.text = normalize_whitespace(cldata.get("abstract", ''))

    #After the collection's did element.
    colbioghistelement = etree.SubElement(archdescelement, "bioghist")
    colbioghistheadelement = etree.SubElement(colbioghistelement, "head")
    colbioghistheadelement.text = "Biographical/Historical Note"

    bioghistlines = cldata.get("bioHistNote", '').splitlines()
    for line in bioghistlines:
        pelement = etree.SubElement(colbioghistelement, "p")
        pelement.text = ' '.join(line.split())

    #descriptive descgrp
    coldescgrpdescriptiveelement = etree.SubElement(archdescelement, "descgrp", {"type":"descriptive"})

    colcollectioninformationelement = etree.SubElement(coldescgrpdescriptiveelement, "head")
    colcollectioninformationelement.text = "Collection information"

    colscopenoteelement = etree.SubElement(coldescgrpdescriptiveelement, "scopecontent")

    colscopelines = cldata.get("scopeNote", '').splitlines()
    for line in colscopelines:
        pelement = etree.SubElement(colscopenoteelement, "p")
        pelement.text = ' '.join(line.split())

    coluserestrictelement = etree.SubElement(coldescgrpdescriptiveelement, "userestrict")
    coluserestrictlines = cldata.get("conditionsUse", '').splitlines()
    for line in coluserestrictlines:
        pelement = etree.SubElement(coluserestrictelement, "p")
        pelement.text = ' '.join(line.split())

    colaccessrestrictelement = etree.SubElement(coldescgrpdescriptiveelement, "accessrestrict")
    colaccessrestrictlines = cldata.get("conditionsAccess", '').splitlines()
    for line in colaccessrestrictlines:
        pelement = etree.SubElement(colaccessrestrictelement, "p")
        pelement.text = ' '.join(line.split())

    colpreferciteelement = etree.SubElement(coldescgrpdescriptiveelement, "prefercite")
    colprefercitelines = cldata.get("preferredCitation", '').splitlines()
    for line in colprefercitelines:
        pelement = etree.SubElement(colpreferciteelement, "p")
        pelement.text = ' '.join(line.split())

    colarrangementelement = etree.SubElement(coldescgrpdescriptiveelement, "arrangement")
    arrangementnotepelement = etree.SubElement(colarrangementelement, "p")
    arrangementnotepelement.text = normalize_whitespace(cldata.get("arrangementNote", ''))
    conlyserieslist = etree.SubElement(colarrangementelement, "list")

    #administrative descgrp
    coldescgrpadministrativeelement = etree.SubElement(archdescelement, "descgrp", {"type":"administrative"})

    administrativeinformationcodes = ['acquisitionInformation',"processingInformation","custodialHistory",'accruals','appraisal']

    for code in administrativeinformationcodes:
        if cldata.get(code,'') != '':
            coladministrativeheadelement = etree.SubElement(coldescgrpadministrativeelement, "head")
            coladministrativeheadelement.text = "Administrative information"
            break

    colacqinfoadminelement = etree.SubElement(coldescgrpadministrativeelement, "acqinfo")
    colacqinfoadminlines = cldata.get("acquisitionInformation", '').splitlines()
    for line in colacqinfoadminlines:
        pelement = etree.SubElement(colacqinfoadminelement, "p")
        pelement.text = ' '.join(line.split())

    colprocessinfoadminelement = etree.SubElement(coldescgrpadministrativeelement, "processinfo")
    colprocessinfoadminlines = cldata.get("processingInformation", '').splitlines()
    for line in colprocessinfoadminlines:
        pelement = etree.SubElement(colprocessinfoadminelement, "p")
        pelement.text = ' '.join(line.split())

    colcustodhistadminelement = etree.SubElement(coldescgrpadministrativeelement, "custodhist")
    colcustodhistadminlines = cldata.get("custodialHistory", '').splitlines()
    for line in colcustodhistadminlines:
        pelement = etree.SubElement(colcustodhistadminelement, "p")
        pelement.text = ' '.join(line.split())

    multilinefield(cldata, coldescgrpadministrativeelement, 'accruals', 'accruals')
    multilinefield(cldata, coldescgrpadministrativeelement, 'appraisal', 'appraisal')

    #descgrp additional
    coldescgrpadditionalelement = etree.SubElement(archdescelement, "descgrp", {"type":"additional"})

    if cldata.get('generalNote','') != '' or cldata.get('relatedMaterials','') != '' or cldata.get('separatedMaterials','') != '' or cldata.get('locationOriginals','') != '' or cldata.get('otherFindingAids','') != '' or cldata.get('otherFormats','') != '':
        coldescgrpadditionalheaderelement = etree.SubElement(coldescgrpadditionalelement, "head")
        coldescgrpadditionalheaderelement.text = "Additional information"

    multilinefield(cldata, coldescgrpadditionalelement, 'generalNote', 'odd')
    multilinefield(cldata, coldescgrpadditionalelement, 'relatedMaterials', 'relatedmaterial')
    multilinefield(cldata, coldescgrpadditionalelement, 'separatedMaterials', 'separatedmaterial')
    multilinefield(cldata, coldescgrpadditionalelement, 'locationOriginals', 'originalsloc')
    multilinefield(cldata, coldescgrpadditionalelement, 'otherFindingAids', 'otherfindaid')
    multilinefield(cldata, coldescgrpadditionalelement, 'otherFormats', 'altformavail')

    #descgrp cataloging
    coldescgrpcatalogingelement = etree.SubElement(archdescelement, "descgrp", {"type":"cataloging"})
    coldescgrpnamescontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")

    coldescgrpcontrolaccessnamesheadelement = etree.SubElement(coldescgrpnamescontrolaccesselement, "head")
    if cldata.get("addedEntryPersonLC", '') != '' or cldata.get("addedEntryPersonLocal", '') != '' or cldata.get("addedEntryCorporateLC", '') != '' or cldata.get("addedEntryCorporateLocal", '') != '':
        coldescgrpcontrolaccessnamesheadelement.text = "Names"

    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryPersonLC', 'persname', {'source':'lcnaf'})
    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryPersonLocal', 'persname', {'source':'local'})

    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryCorporateLC', 'corpname', {'source':'lcnaf'})
    repeatingsubjectfield(coldescgrpnamescontrolaccesselement, cldata, 'addedEntryCorporateLocal', 'corpname', {'source':'local'})

    coldescgrpsubjectscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccesssubjectheadelement = etree.SubElement(coldescgrpsubjectscontrolaccesselement, "head")
    if cldata.get("addedEntrySubjectLC", '') != '' or cldata.get("addedEntrySubjectLocal", '') != '' or cldata.get("addedEntrySubjectFAST", '') != '' or cldata.get("addedEntryGeographicLC", '') != '' or cldata.get("addedEntryGeographicLocal", '') != '':
        coldescgrpcontrolaccesssubjectheadelement.text = "Subjects"

    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntrySubjectLC', 'subject', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntrySubjectLocal', 'subject', {'source':'local'})
    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntrySubjectFAST', 'subject', {'source':'fast'})

    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntryGeographicLC', 'geogname', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpsubjectscontrolaccesselement, cldata, 'addedEntryGeographicLocal', 'geogname', {'source':'local'})

    coldescgrpoccupationscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccessoccupationheadelement = etree.SubElement(coldescgrpoccupationscontrolaccesselement, "head")
    if cldata.get("addedEntryOccupationLC", '') != '' or cldata.get("addedEntryOccupationLocal", '') != '':
        coldescgrpcontrolaccessoccupationheadelement.text = "Occupations"

    repeatingsubjectfield(coldescgrpoccupationscontrolaccesselement, cldata, 'addedEntryOccupationLC', 'occupation', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpoccupationscontrolaccesselement, cldata, 'addedEntryOccupationLocal', 'occupation', {'source':'local'})

    coldescgrpmaterialscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccessgenreheadelement = etree.SubElement(coldescgrpmaterialscontrolaccesselement, "head")
    if cldata.get("addedEntryGenreAAT", '') != '' or cldata.get("addedEntryGenreLCSH", '') != '' or cldata.get("addedEntryGenreTGM", '') != '' or cldata.get("addedEntryGenreRBGENR", '') != '' or cldata.get("addedEntryGenreLocal", '') != '':
        coldescgrpcontrolaccessgenreheadelement.text = "Types of Materials"

    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreAAT', 'genreform', {'source':'aat'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreLCSH', 'genreform', {'source':'lcsh'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreTGM', 'genreform', {'source':'tgm'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreRBGENR', 'genreform', {'source':'rbgenr'})
    repeatingsubjectfield(coldescgrpmaterialscontrolaccesselement, cldata, 'addedEntryGenreLocal', 'genreform', {'source':'local'})

    coldescgrptitlescontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccesstitleheadelement = etree.SubElement(coldescgrptitlescontrolaccesselement, "head")
    if cldata.get("addedEntryTitle", '') != '':
        coldescgrpcontrolaccesstitleheadelement.text = "Titles"

    repeatingsubjectfield(coldescgrptitlescontrolaccesselement, cldata, 'addedEntryTitle', 'title', {})

    coldescgrpriamcoscontrolaccesselement = etree.SubElement(coldescgrpcatalogingelement, "controlaccess")
    coldescgrpcontrolaccessriamcoheadelement = etree.SubElement(coldescgrpriamcoscontrolaccesselement, "head")
    if cldata.get("RIAMCOBrowsingTerm", '') != '':
        coldescgrpcontrolaccessriamcoheadelement.text = "RIAMCO Browsing Term"

    repeatingsubjectfield(coldescgrpriamcoscontrolaccesselement, cldata, 'RIAMCOBrowsingTerm', 'subject', {'altrender':'nodisplay','source':'riamco'})

    #Create the container list.
    dscelement = etree.SubElement(archdescelement, "dsc", {"type":"combined"})
    ctelement = dscelement
    crecordgroupelement = dscelement
    csubgroupelement = dscelement
    cserieselement = dscelement
    csubserieselement = dscelement
    csubsubserieselement = dscelement
    cfileelement = dscelement
    rowindex = 2
    onlySeriesRows = True

    #Set up in the case that there is only series and subseries.
    arrangementnotepforlists = etree.SubElement(colarrangementelement, "p")
    cserieslist = etree.SubElement(arrangementnotepforlists, "list")
    csubserieslist = etree.SubElement(cserieslist, "list")

    for row in csvdata:
        if row.get("recordgroupTitle", '') != '' or row.get("subgroupTitle", '') != '' or row.get("subSeriesTitle", '') != '':
            onlySeriesRows = False
            break

    if onlySeriesRows:
        arrangementnoteseriespelement = etree.SubElement(colarrangementelement, "p")
        cserieslist = etree.SubElement(arrangementnoteseriespelement, "list")
        print("Only series rows is True.", file=sys.stderr)

    for row in csvdata:

        if row.get('Ignore', '') != '':
            continue
        if row.get("recordgroupTitle", '') != "":
            serieselement = etree.SubElement(dscelement, "c", {"id":("c"+str(rowindex)), "level":"recordgrp"})
            crecordgroupelement = serieselement
            csubgroupelement = serieselement
            cserieselement = serieselement
            csubserieselement = serieselement
            cfileelement = serieselement
            ctelement = serieselement

            didelement = etree.SubElement(ctelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("recordgroupTitle", '').split())
            cunittitle = titleelement

            seriesIDelement = etree.SubElement(cdid, "unitid", {"type":"recordgrp"})
            cseriesID = "Record Group " + ' '.join(str(row.get("recordgroupID", '')).split()).replace('.0','')
            seriesIDelement.text = cseriesID

            #Add the series to the Arrangement Note.
            arrangementnotepaddition = etree.SubElement(colarrangementelement, "p")
            arrangementnoteemphaddition = etree.SubElement(arrangementnotepaddition, "emph", {"render":"bold"})
            arrangementnoteemphaddition.text = cseriesID + ". " + titleelement.text

            arrangementnotepforlists = etree.SubElement(colarrangementelement, "p")
            cserieslist = etree.SubElement(arrangementnotepforlists, "list")
            csubserieslist = etree.SubElement(cserieslist, "list")
        if row.get("subgroupTitle", '') != "":
            serieselement = etree.SubElement(crecordgroupelement, "c", {"id":("c"+str(rowindex)), "level":"subgrp"})
            csubgroupelement = serieselement
            cserieselement = serieselement
            csubserieselement = serieselement
            cfileelement = serieselement
            ctelement = serieselement

            didelement = etree.SubElement(ctelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("subgroupTitle", '').split())
            cunittitle = titleelement

            seriesIDelement = etree.SubElement(cdid, "unitid", {"type":"subgrp"})
            cseriesID = "Subgroup " + ' '.join(str(row.get("subgroupID", '')).split()).replace('.0','')
            seriesIDelement.text = cseriesID

            #Add the series to the Arrangement Note.
            arrangementnotepaddition = etree.SubElement(colarrangementelement, "p")
            arrangementnotepaddition.text = cseriesID + ". " + titleelement.text

            arrangementnotepforlists = etree.SubElement(colarrangementelement, "p")
            cserieslist = etree.SubElement(arrangementnotepforlists, "list")
            csubserieslist = etree.SubElement(cserieslist, "list")
        #Create a top level series element if the series cell is not blank.
        if row.get("seriesTitle", '') != "":
            serieselement = etree.SubElement(csubgroupelement, "c", {"id":("c"+str(rowindex)), "level":"series"})
            cserieselement = serieselement
            csubserieselement = serieselement
            cfileelement = serieselement
            ctelement = serieselement

            didelement = etree.SubElement(ctelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("seriesTitle", '').split())
            cunittitle = titleelement

            seriesIDelement = etree.SubElement(cdid, "unitid", {"type":"series"})
            cseriesID = "Series " + ' '.join(str(row.get("seriesID", '')).split()).replace('.0','')
            seriesIDelement.text = cseriesID

            seriesarrangementelement = etree.SubElement(cserieslist, "item")
            seriesarrangementelement.text = cseriesID + ". " + titleelement.text

            csubserieslist = etree.SubElement(seriesarrangementelement, "list")

        #Create a top level subseries element if the subseries cell is not blank.
        elif row.get("subSeriesTitle", '') != "":
            subserieselement = etree.SubElement(cserieselement, "c", {"id":("c"+str(rowindex)), "level":"subseries"})
            csubserieselement = subserieselement
            cfileelement = subserieselement
            ctelement = subserieselement

            didelement = etree.SubElement(subserieselement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("subSeriesTitle", '').split())
            cunittitle = titleelement

            subseriesIDelement = etree.SubElement(cdid, "unitid", {"type":"subseries"})
            subseriesIDelement.text = cseriesID + ". Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','')

            subseriesarrangementelement = etree.SubElement(csubserieslist, "item")
            subseriesarrangementelement.text = "Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','') + ". " + titleelement.text
        #Create a top level file element if the title cell is not blank.
        elif row.get("fileTitle", '') != "":
            fileement = etree.SubElement(csubserieselement, "c", {"id":("c"+str(rowindex)), "level":"file"})
            cfileelement = fileement
            ctelement = fileement

            didelement = etree.SubElement(fileement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("fileTitle", '').split())
            cunittitle = titleelement

        #Create a top level item element if the title cell is not blank.
        elif row.get("itemTitle", '') != "":
            itemelement = etree.SubElement(cfileelement, "c", {"id":("c"+str(rowindex)), "level":"item"})
            ctelement = itemelement

            didelement = etree.SubElement(itemelement, "did")
            cdid = didelement

            titleelement = etree.SubElement(cdid, "unittitle")
            titleelement.text = ' '.join(row.get("itemTitle", '').split())
            cunittitle = titleelement
            #Needs attention: Item ID.
            #itemIDelement = etree.SubElement(didelement, "unitid", {"type":"subseries"})
            #itemIDelement.text = cseriesID + ". Subseries " + ' '.join(str(row.get("subSeriesID", '')).split()).replace('.0','')

        #container
        shelfLocatorstring = ""
        barcodestring = ''

        #if row.get("barcode", '') != '':
        #    floatbarcode = float(row.get("barcode", ''))

        if row.get("barcode", '') != '':
            barcodestring = ' [' + normalize_whitespace(row.get("barcode", '')).replace(".0", "") + ']'

        for i in 1,2,3:
            colname = "shelfLocator" + str(i)
            cell = row.get(colname, '')
            if cell:
                shelflocatorattrs = {"type":' '.join(cell.split()).lower().replace(' ', '_'), "label": normalize_whitespace(cell.title()) + barcodestring}
                shelflocatorelement = etree.SubElement(cdid, "container", shelflocatorattrs)
                shelflocatorelement.text = ' '.join(str(row.get(colname + "ID", '')).split()).replace('.0','')

        #dates
        #Test for a YYYY - YYYY and remove dates if so.
        match = re.search(u"(\d{4}\s-\s\d{4})", row.get("dateText", ''))

        inclusivedatetext= normalize_whitespace(row.get("dateText", '')).replace('.0','') #' '.join(row.get("dateText", '').split()).replace('.0','')
        if match:
            inclusivedatetext = inclusivedatetext.replace(' ','')
        inclusivedatestart = normalize_whitespace(row.get("dateStart", '')).replace('.0','') #' '.join(row.get("dateStart", '').split()).replace('.0','')
        inclusivedateend = normalize_whitespace(row.get("dateEnd", '')).replace('.0','') #' '.join(row.get("dateEnd", '').split()).replace('.0','')

        bulkdatestart = normalize_whitespace(row.get("dateBulkStart", '')).replace('.0','') # ' '.join(row.get("dateBulkStart", '').split()).replace('.0','')
        bulkdateend = normalize_whitespace(row.get("dateBulkEnd", '')).replace('.0','') #' '.join(row.get("dateBulkEnd", '').split()).replace('.0','')

        unitdateinclusiveattributes = {"type":"inclusive"}
        if inclusivedatestart != '' and inclusivedateend != '':
            unitdateinclusiveattributes["normal"] = inclusivedatestart +"/"+inclusivedateend
        if row.get("dateQualifier", '') != "":
            unitdateinclusiveattributes["certainty"] = normalize_whitespace(row.get("dateQualifier", ''))
        unitdatebulkattributes = {"type":"bulk","normal":(bulkdatestart +"/"+bulkdateend)}

        inclusivedateelement = etree.SubElement(cdid, "unitdate", unitdateinclusiveattributes)
        if inclusivedatetext == '' and (inclusivedatestart.count != '' or inclusivedateend != ''):
            if inclusivedatestart != inclusivedateend:
                inclusivedatetext = inclusivedatestart + ' - ' + inclusivedateend
                inclusivedatetext = inclusivedatetext.lstrip(' - ').rstrip(' - ')
            else:
                inclusivedatetext = inclusivedatestart
        inclusivedateelement.text = inclusivedatetext

        if bulkdatestart != "" or bulkdateend != "":
            bulkdateelement = etree.SubElement(cdid, "unitdate", unitdatebulkattributes)
            bulkdateelement.text = "(bulk " + bulkdatestart + "-" + bulkdateend + ")"

        #physicalDescription
        physicalDescriptionelement = etree.SubElement(cdid, "physdesc")

        #extent and genre
        extentquantityphysdescelement = etree.SubElement(cdid, "physdesc", {"altrender":"whole"})
        extentQuantityelement = etree.SubElement(extentquantityphysdescelement, "extent", {"altrender":"materialtype spaceoccupied"})
        extentQuantityelement.text = ' '.join(row.get("extentQuantity", '').split())
        containerSummaryelement = etree.SubElement(extentquantityphysdescelement, "extent",{"altrender":"carrier"})
        containerSummaryelement.text = normalize_whitespace(row.get("containerSummary",''))

        extentsizephysdescelement = etree.SubElement(cdid, "physdesc")
        extentSizeelement = etree.SubElement(extentsizephysdescelement, "dimensions")
        extentSizeelement.text = ' '.join(row.get("extentSize", '').split())

        extentSpeedelement = etree.SubElement(extentsizephysdescelement, "dimensions")
        extentSpeedelement.text = ' '.join(row.get("extentSpeed", '').split())

        genreformphyscdescelement = etree.SubElement(cdid, "physdesc")
        genreSources = {
            'genreAAT': "aat",
            'genreLCSH': "lcsh",
            'genreLocal': "local",
            'genreRBGENR': "rbgenr"
        }

        for genre, source in genreSources.items():
            repeatingsubjectfield(
                parentelement = genreformphyscdescelement,
                refdict = row,
                originalfieldname = genre,
                eadfieldname = 'genreform',
                eadattributes = {"source":source}
            )

        #materialspec
        formelement = etree.SubElement(cdid, "materialspec")
        formelement.text = ' '.join(row.get("form", '').split())

        #language
        langmaterialelement = etree.SubElement(cdid, "langmaterial")
        langmaterialelementsplitchar = getSplitCharacter(row.get("language", ''))

        for language in row.get("language", '').split(langmaterialelementsplitchar):
            if language == "":
                continue
            if len(normalize_whitespace(language)) < 4:
                language = langcodeopp.get(language, '')
            if normalize_whitespace(language) in langcode:
                langusagelangelement = etree.SubElement(langmaterialelement, "language", {"langcode":langcode[normalize_whitespace(language)], "scriptcode":scriptcode[normalize_whitespace(language)]})
                langusagelangelement.text =  normalize_whitespace(language)

                if scriptcode == "N/A":
                    langissue = True
            else:
                langusagelangelement = etree.SubElement(langmaterialelement, "language", {"langcode":"***", "scriptcode":"***"})
                langusagelangelement.text =  normalize_whitespace(language)

                langissue = True

        #Create origination and controlaccess element.

        originationelement = etree.SubElement(cdid, "origination")

        #note
        notescopeelement = etree.SubElement(ctelement, "scopecontent")
        scopelines = row.get("noteScope", '').splitlines()
        for line in scopelines:
            pelement = etree.SubElement(notescopeelement, "p")
            pelement.text = ' '.join(line.split())

        notebioghistelement = etree.SubElement(ctelement, "bioghist")
        bioghistlines = row.get("noteHistorical", '').splitlines()
        for line in bioghistlines:
            pelement = etree.SubElement(notebioghistelement, "p")
            pelement.text = ' '.join(line.split())

        acqinfoelement = etree.SubElement(ctelement, "acqinfo")
        noteaccessionpelement = etree.SubElement(acqinfoelement, "p")
        noteaccessionelement = etree.SubElement(noteaccessionpelement, "num", {"type":"accession"})
        noteaccessionelement.text = ' '.join(row.get("noteAccession", '').split())

        #controlaccess and name fields
        ccontrolaccess = etree.SubElement(ctelement, "controlaccess")

        namefields = [
            {"field":'namePersonCreatorLC', "nametype":'persname', "role":'creator', "source":'naf'},
            {"field":'namePersonCreatorLocal', "nametype":'persname', "role":'creator', "source":'local'},
            {"field":'namePersonCreatorFAST', "nametype":'persname', "role":'creator', "source":'fast'},
            {"field":'nameCorpCreatorLC', "nametype":'corpname', "role":'creator', "source":'naf'},
            {"field":'nameCorpCreatorLocal', "nametype":'corpname', "role":'creator', "source":'local'},
            {"field":'nameCorpCreatorFAST', "nametype":'corpname', "role":'creator', "source":'fast'},
            {"field":'namePersonOtherLC', "nametype":'persname', "role":'', "source":'naf'},
            {"field":'namePersonOtherLocal', "nametype":'persname', "role":'', "source":'local'},
            {"field":'namePersonOtherFAST', "nametype":'persname', "role":'', "source":'fast'}
        ]

        for nf in namefields:
            repeatingNameField(
                parentElement=originationelement,
                elementName=nf['nametype'],
                rowString=row.get(nf['field'], ''),
                assignedRole=nf['role'],
                source=nf['source']
            )

        notegeneralelement = etree.SubElement(ctelement, "odd")
        scopelines = row.get("noteGeneral", '').splitlines()
        for line in scopelines:
            pelement = etree.SubElement(notegeneralelement, "p")
            pelement.text = ' '.join(line.split())

        notealtformelement = etree.SubElement(ctelement, "altformavail")
        altformlines = row.get("locationCopies", '').splitlines()
        for line in altformlines:
            pelement = etree.SubElement(notealtformelement, "p")
            pelement.text = ' '.join(line.split())

        #geogname
        geognameelement = etree.SubElement(cunittitle, "geogname")
        geognameelement.text = ' '.join(row.get("place", '').split())

        #subject
        subjectRows = {
            'subjectNamesLC': {'type': 'persname', 'source': 'naf'},
            'subjectNamesLocal': {'type': 'persname', 'source': 'local'},
            'subjectNamesFAST': {'type': 'persname', 'source': 'fast'},

            'subjectCorpLC': {'type': 'corpname', 'source': 'naf'},
            'subjectCorpLocal': {'type': 'corpname', 'source': 'local'},
            'subjectCorpFAST': {'type': 'corpname', 'source': 'fast'},

            'subjectTopicsLC': {'type': 'subject', 'source': 'naf'},
            'subjectTopicsLocal': {'type': 'subject', 'source': 'local'},
            'subjectTopicsFAST': {'type': 'subject', 'source': 'fast'},

            'subjectGeoLC': {'type': 'geogname', 'source': 'lcsh'},
            'subjectGeoFAST': {'type': 'geogname', 'source': 'fast'},

            'subjectTemporalLC': {'type': 'subject', 'source': 'lcsh'},
            'subjectTemporalFAST': {'type': 'subject', 'source': 'fast'},

            'subjectTitleLC': {'type': 'title', 'source': 'lcsh'},
            'subjectTitleFAST': {'type': 'title', 'source': 'fast'}
        }

        for subjectRow in subjectRows:
            rowtype = subjectRows[subjectRow]['type']
            source = subjectRows[subjectRow]['source']
            repeatingsubjectfield(
                parentelement = ccontrolaccess,
                refdict = row,
                originalfieldname = subjectRow,
                eadfieldname = rowtype,
                eadattributes = {"source": source}
            )


        #dao fields
        if row.get("identifierBDR", ''):
            # make_dao_element(ctelement,cunittitle, "MODS_ID", 'bdr'+ xmltext(row.get("identifierBDR", '')).lstrip('bdr').replace(':',''))
            make_dao_element(ctelement,cunittitle, "BDR_PID", 'bdr:'+ ' '.join(row.get("identifierBDR", '').split()).lstrip('bdr').replace(':',''))

        if row.get("identifierNormalized", ''):
            make_dao_element(ctelement,cunittitle,"NORMALIZEDFILE_ID", normalize_whitespace(row.get("identifierNormalized", '')).lstrip('bdr'))

        if row.get("identifierWebArchive", ''):
            make_dao_element(ctelement,cunittitle,"WEBARCHIVEURL",normalize_whitespace(row.get("identifierWebArchive", '')).lstrip('bdr'))

        if row.get("identifierFileName", ''):
            make_dao_element(ctelement,cunittitle,"BDR_PID",normalize_whitespace(row.get("identifierFileName", '')))

        rowindex = rowindex + 1

    #Create the collection-level data.
    eadidattributes = {"countrycode":"US", "mainagencycode":"US-" + normalize_whitespace(cldata.get("MARCRepositoryCode", '')), "identifier":normalize_whitespace(cldata.get("callNumber", '')).lower()+'.xml'}
    eadidelement = etree.SubElement(eadheaderelement, "eadid", eadidattributes)
    eadidelement.text = "US-"+normalize_whitespace(cldata.get("MARCRepositoryCode", ''))+"-"+normalize_whitespace(cldata.get("callNumber", '')).lower()

    filedescelement = etree.SubElement(eadheaderelement, "filedesc")

    #titlestmt
    titlestmtelement = etree.SubElement(filedescelement, "titlestmt")
    titleproperelement = etree.SubElement(titlestmtelement, "titleproper")
    titleproperelement.text = "Guide to the " + normalize_whitespace(cldata.get("title", ''))

    inclusivedateattributes = {"type":"inclusive", "era":"ce","calendar":"gregorian","normal":normalize_whitespace(cldata.get("inclusiveDates", '')).replace('-','/').replace(' ','')}
    inclusivedateelement = etree.SubElement(titleproperelement, "date", inclusivedateattributes)
    inclusivedateelement.text = normalize_whitespace(cldata.get("inclusiveDates", '').replace(' ',''))

    bulkdateattributes = {"type":"bulk", "era":"ce","calendar":"gregorian","normal":normalize_whitespace(cldata.get("bulkDates", '')).replace('-','/').replace(' ','')}
    bulkdateelement = etree.SubElement(titleproperelement, "date", bulkdateattributes)
    bulkdateelement.text = "(bulk "+ normalize_whitespace(cldata.get("bulkDates", '')).replace(' ','') + ")"

    authorelement = etree.SubElement(titlestmtelement, "author").text = "Finding aid prepared by " + normalize_whitespace(cldata.get("author", ''))

    sponsorelement = etree.SubElement(titlestmtelement, "sponsor").text = normalize_whitespace(cldata.get("sponsor", ''))

    #publicationstmt
    publicationstmtelement = etree.SubElement(filedescelement, "publicationstmt")

    publisherelement = etree.SubElement(publicationstmtelement, "publisher")
    publisherelement.text = normalize_whitespace(cldata.get("publisher", ''))

    pubaddresselement = etree.SubElement(publicationstmtelement, "address")

    pubaddresslines = cldata.get("address", '').splitlines()
    for line in pubaddresslines:
        addresselement = etree.SubElement(pubaddresselement, "addressline")
        addresselement.text = normalize_whitespace(line)

    creationdateelement = etree.SubElement(publicationstmtelement, "date", {"era":"ce","calendar":"gregorian", "normal":normalize_whitespace(cldata.get("creationDate", ''))[:4], "type":"publication"})
    creationdateelement.text = normalize_whitespace(cldata.get("creationDate", '').replace('.0',''))

    #profiledesc
    profiledescelement = etree.SubElement(eadheaderelement, "profiledesc")

    creationelement = etree.SubElement(profiledescelement, "creation")
    creationelement.text = "This finding aid was produced using the RIAMCO EAD spreadsheet, "
    creationdatecreationelement = etree.SubElement(creationelement, "date", {"era":"ce","calendar":"gregorian", "normal":normalize_whitespace(cldata.get("creationDate", ''))[:4], "type":"publication"})
    creationdatecreationelement.text = normalize_whitespace(cldata.get("creationDate", '').replace('.0',''))

    #langusage
    langusageelement = etree.SubElement(profiledescelement, "langusage")

    if normalize_whitespace(cldata.get("findingAidLanguage", '')) in langcode:
        langusagelangelement = etree.SubElement(langusageelement, "language", {"langcode":langcode[normalize_whitespace(cldata.get("findingAidLanguage", ''))], "scriptcode":scriptcode[normalize_whitespace(cldata.get("findingAidLanguage", ''))]})
        langusagelangelement.text = normalize_whitespace(cldata.get("findingAidLanguage", ''))

        if scriptcode == "N/A":
            langissue = True
    else:
        langusagelangelement = etree.SubElement(langusageelement, "language", {"langcode":"***", "scriptcode":"***"})
        langusagelangelement.text = normalize_whitespace(cldata.get("findingAidLanguage", ''))

        langissue = True

    descruleselement = etree.SubElement(profiledescelement, "descrules")
    descruleselement.text = "Finding aid based on Describing Archives: A Content Standard (DACS)"



    # start cleanup
    # remove any element tails
    for element in eadtop.iter():
        element.tail = None

    # remove any line breaks or tabs in element text
        if element.text:
            if '\n' in element.text:
                element.text = element.text.replace('\n', '')
            if '\t' in element.text:
                element.text = element.text.replace('\t', '')

    # remove any remaining whitespace
    parser = etree.XMLParser(remove_blank_text=True, remove_comments=True, recover=True)
    treestring = etree.tostring(eadtop)
    clean = etree.XML(treestring, parser)

    # remove recursively empty nodes
    # found here: https://stackoverflow.com/questions/12694091/python-lxml-how-to-remove-empty-repeated-tags
    def recursively_empty(e):
       if e.text:
           return False
       return all((recursively_empty(c) for c in e.iterchildren()))

    context = etree.iterwalk(clean)
    for action, elem in context:
        parent = elem.getparent()
        if recursively_empty(elem):
            parent.remove(elem)

    # remove nodes with blank attribute
    for element in clean.xpath(".//*[@*='']"):
        element.getparent().remove(element)

    # remove nodes with attribute "null"
    for element in clean.xpath(".//*[@*='null']"):
        element.getparent().remove(element)

    completestring = etree.tostring(clean, pretty_print = True, encoding="unicode")
    completestring = completestring.replace('&lt;','<')
    completestring = completestring.replace('&gt;','>')

    finalfilename = eadidelement.text + '.xml'

    with open(CACHEDIR + id + "/" + finalfilename, 'w+') as f:
        f.write("<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n")
        f.write(completestring)
        f.close()


    returndict = {}

    returndict["filename"] = finalfilename
    returndict["error"] = False
    returndict["allrecords"] = completestring

    with open(CACHEDIR + id + "/" + finalfilename, 'rb') as f:
        return(f.read(), returndict)

    if langissue:
        #print("")
        print("*Language Field Error*\n", file=sys.stderr)
        print("There were one or more issues with language fields in your spreadsheet. Please check your spelling in all language fields. You may also manually correct your XML file, consult the SupportedLanguages.xlsx file in the data folder for supported languages, and/or adjust the SupportedLanguages.xlsx spreadsheet to suit your project.\n", file=sys.stderr)

        print('\n\n', file=sys.stderr)

    #errorfile.close()
    print("***Operation Complete***\n", file=sys.stderr)
    #rint("Your EAD file was written to folder " + output_path + ".\n", file=sys.stderr)

    print('\n\n', file=sys.stderr)

    print( u"   \u2606 \u2606 \u2606", file=sys.stderr)
    print('\n\n\n', file=sys.stderr)

